import csv
import json
import os
import platform
import tempfile
import threading
from itertools import islice

import openpyxl
import pandas as pd

import Common.Globals as Globals


def checkIfCurrentThreadStopped():
    """Check if the current thread has been asked to stop/abort.
    This function is defined here to avoid circular dependency with Utility.Resource"""
    isAbortSet = False
    if hasattr(threading.current_thread(), "abort"):
        isAbortSet = threading.current_thread().abort.is_set()
    elif hasattr(threading.current_thread(), "isStopped"):
        isAbortSet = threading.current_thread().isStopped()
    elif hasattr(threading.current_thread(), "stopCurrentTask"):
        isAbortSet = threading.current_thread().stopCurrentTask.is_set()
    return isAbortSet


def read_from_file(filePath, mode="r") -> list:
    content = None
    try:
        with open(filePath, mode) as file:
            content = file.read()
    except (FileNotFoundError, PermissionError, IOError) as e:
        print(f"Error reading file {filePath}: {e}")
        content = None
    return content


def read_json_file(filePath) -> dict:
    content = {}
    try:
        with open(filePath, "r") as file:
            content = json.load(file)
            if content is None:
                content = {}
    except (FileNotFoundError, json.JSONDecodeError, Exception):
        content = {}
    return content


def write_json_file(filePath, data: dict):
    with open(filePath, "w") as outfile:
        if data:
            json.dump(data, outfile)


def write_content_to_file(filePath, data, mode="w", encoding="utf-8") -> None:
    if "b" in mode:
        encoding = None
    with open(filePath, mode, encoding=encoding) as file:
        if type(data) is list:
            for line in data:
                file.write(line)
        elif data:
            file.write(data)


def read_data_from_csv_as_dict(filePath: str) -> list:
    fileData = None
    if filePath.endswith("csv"):
        with open(filePath, "r") as csvFile:
            csv_reader = csv.DictReader(csvFile)
            fileData = list(csv_reader)
    return fileData


def write_data_to_csv(filePath: str, data, mode="w", encoding="utf-8", newline="") -> None:
    try:
        with open(filePath, mode, newline=newline, encoding=encoding) as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_NONNUMERIC)
            if type(data) is dict:
                writer.writerow(list(data.values()))
            elif any(isinstance(el, list) for el in data):
                writer.writerows(data)
            elif type(data) is list:
                writer.writerow(data)
            else:
                if data:
                    writer.writerows(data)
    except Exception as e:
        print(e)
        raise e


def getToolDataPath():
    basePath = "%s/EsperApiTool/" % tempfile.gettempdir().replace("Local", "Roaming").replace("Temp", "")

    if platform.system() != "Windows":
        # Use Application Support directory on macOS (no permission prompt required)
        basePath = os.path.expanduser("~/Library/Application Support/EsperApiTool/")
        
        # Migrate existing data from old Desktop location (one-time migration)
        _migrateFromDesktopToAppSupport(basePath)

    return basePath


def _migrateFromDesktopToAppSupport(newBasePath):
    """
    Migrate existing preferences and data from Desktop to Application Support.
    This is a one-time migration for users upgrading from older versions.
    """
    if platform.system() == "Windows":
        return  # Only needed for macOS
    
    oldBasePath = os.path.expanduser("~/Desktop/EsperApiTool/")
    
    # Skip if old location doesn't exist or new location already has files
    if not os.path.exists(oldBasePath):
        return
    if os.path.exists(newBasePath) and os.listdir(newBasePath):
        return  # New location already populated, skip migration
    
    try:
        # Create new directory if it doesn't exist
        if not os.path.exists(newBasePath):
            os.makedirs(newBasePath, exist_ok=True)
        
        # Move files from old to new location
        import shutil
        for item in os.listdir(oldBasePath):
            oldPath = os.path.join(oldBasePath, item)
            newPath = os.path.join(newBasePath, item)
            
            if os.path.isfile(oldPath):
                shutil.copy2(oldPath, newPath)  # Copy with metadata
            elif os.path.isdir(oldPath):
                shutil.copytree(oldPath, newPath, dirs_exist_ok=True)
        
        # Optionally remove old directory (commented out for safety)
        # shutil.rmtree(oldBasePath)
        
    except (OSError, IOError, PermissionError) as e:
        # Silently fail - user can manually migrate if needed
        print(f"Note: Could not migrate data from Desktop: {e}")


def read_excel_via_openpyxl(path: str, readAnySheet=False) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    df = None
    rows = []
    headers = None
    for sheet in workbook.sheetnames:
        if "Device & Network" in sheet or "Device and Network" in sheet or "Device" in sheet or readAnySheet:
            # Select the worksheet
            worksheet = workbook[sheet]
            # Extract the data
            for row in worksheet.iter_rows(values_only=True):
                if not headers:
                    headers = row
                # Try to avoid shape mismatch
                if len(row) == len(headers):
                    rows.append(row)
    df = None
    if len(rows) > 0:
        df = pd.DataFrame(rows[1:], columns=rows[0])
    else:
        df = pd.DataFrame()
    return df


def read_csv_via_pandas(path: str) -> pd.DataFrame:
    data = None
    try:
        data = pd.read_csv(path, sep=",", header=0, keep_default_na=False, chunksize=1000)
    except (UnicodeDecodeError, pd.errors.ParserError) as e:
        try:
            # Try to decode ANSI encoded CSV files
            data = pd.read_csv(
                path,
                sep=",",
                header=0,
                keep_default_na=False,
                chunksize=1000,
                encoding="mbcs",
            )
        except Exception as e:
            print(e)
            raise e
    return pd.concat(data, ignore_index=True)


def save_excel_pandas_xlxswriter(path, df_dict: dict):
    # Check abort before starting
    if checkIfCurrentThreadStopped():
        return
    
    if len(df_dict) <= Globals.MAX_NUMBER_OF_SHEETS_PER_FILE:
        # Optimize xlsxwriter with engine options for better performance
        writer = pd.ExcelWriter(
            path,
            engine="xlsxwriter",
            engine_kwargs={
                'options': {
                    'strings_to_numbers': False,  # Skip automatic number conversion
                    'strings_to_formulas': False,  # Don't parse formulas
                    'strings_to_urls': False,  # Don't convert URLs
                    'constant_memory': True,  # Use constant memory mode for large files
                }
            }
        )

        sheetNames = []
        try:
            for sheet, df in df_dict.items():
                # Check abort before processing each sheet
                if checkIfCurrentThreadStopped():
                    # Close writer if abort is detected
                    if hasattr(writer, "close"):
                        writer.close()
                    return
                
                sheetNames.append(sheet)
                # Optimize: Disable automatic date conversion for better performance
                df.to_excel(
                    writer, 
                    sheet_name=sheet, 
                    index=False,
                    freeze_panes=None,  # Disable freeze panes for faster write
                )

            # Auto adjust column width
            for s in sheetNames:
                # Check abort before adjusting each sheet's column width
                if checkIfCurrentThreadStopped():
                    break
                    
                worksheet = writer.sheets[s]
                if hasattr(worksheet, "autofit"):
                    worksheet.autofit()
                else:
                    # Get the dataframe for this sheet
                    sheet_df = df_dict[s]
                    for idx, col in enumerate(sheet_df.columns):  # loop through all columns
                        # Check abort during column width adjustment
                        if checkIfCurrentThreadStopped():
                            break
                        
                        series = sheet_df[col]
                        num_rows = len(series)
                        
                        if num_rows == 0:
                            max_len = len(str(series.name)) + 1
                        elif num_rows <= 5000:
                            # For smaller datasets, scan all rows for accuracy
                            max_len = max(
                                series.astype(str).str.len().max(),
                                len(str(series.name))
                            ) + 1
                        else:
                            # For large datasets, use stratified sampling from beginning, middle, and end
                            # Sample ~5000 rows distributed across the dataset
                            sample_size = 5000
                            step = num_rows // (sample_size // 3)  # Divide samples into 3 groups
                            
                            # Get indices for beginning, middle, and end sections
                            begin_indices = list(range(0, min(sample_size // 3, num_rows)))
                            middle_start = (num_rows // 2) - (sample_size // 6)
                            middle_indices = list(range(max(0, middle_start), min(middle_start + sample_size // 3, num_rows)))
                            end_indices = list(range(max(0, num_rows - sample_size // 3), num_rows))
                            
                            # Combine and get sample
                            sample_indices = begin_indices + middle_indices + end_indices
                            sample = series.iloc[sample_indices]
                            
                            # Calculate max length from sample
                            max_len = max(
                                sample.astype(str).str.len().max(),
                                len(str(series.name))
                            ) + 1
                        
                        # Cap maximum column width to prevent extremely wide columns
                        max_len = min(max_len, 100)
                        worksheet.set_column(idx, idx, max_len)  # set column width
        except Exception as e:
            pass
        finally:
            if hasattr(writer, "save"):
                writer.save()
            if hasattr(writer, "close"):
                writer.close()
    else:
        split_dict_list = list(__split_dict_into_chunks__(df_dict, Globals.MAX_NUMBER_OF_SHEETS_PER_FILE))
        for i in range(0, len(split_dict_list)):
            # Check abort before processing each split file
            if checkIfCurrentThreadStopped():
                return
                
            if i == 0:
                path = path[:-5] + "_{}.xlsx".format(i)
            else:
                path = path[:-7] + "_{}.xlsx".format(i)
            save_excel_pandas_xlxswriter(path, split_dict_list[i])


def __split_dict_into_chunks__(data, size):
    it = iter(data)
    for _ in range(0, len(data), size):
        yield {k: data[k] for k in islice(it, size)}


def save_csv_pandas(path, df):
    # Check abort before starting the save operation
    # Note: pandas to_csv is a blocking operation that can't be interrupted mid-write,
    # but checking here allows the thread to exit before starting a long write operation
    if checkIfCurrentThreadStopped():
        return
    
    df.to_csv(
        path,
        sep=",",
        index=False,
        encoding="utf-8",
        quoting=csv.QUOTE_NONNUMERIC,
    )
