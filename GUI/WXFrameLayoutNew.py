#!/usr/bin/env python

import ast
import csv
import json
import math
import os.path
import platform
import sys
import tempfile
import threading
import time
from datetime import datetime
from pathlib import Path
import openpyxl

import pandas as pd
import wx
import wx.adv as wxadv
import xlsxwriter
from wx.core import TextEntryDialog

import Common.ApiTracker as ApiTracker
import Common.Globals as Globals
import GUI.EnhancedStatusBar as ESB
import Utility.API.EsperTemplateUtil as templateUtil
from Utility.API.WidgetUtility import setWidget
from Utility.API.AuditPosting import AuditPosting
import Utility.EventUtility as eventUtil
from Utility.FileUtility import (
    getToolDataPath,
    read_data_from_csv,
    read_data_from_csv_as_dict,
    read_json_file,
    write_data_to_csv,
    write_json_file,
)
from Utility.GridUtilities import addColToGridRow
import Utility.Threading.wxThread as wxThread

from Common.decorator import api_tool_decorator
from Common.enum import Color, GeneralActions, GridActions
from Common.SleepInhibitor import SleepInhibitor
from GUI.ConfigureWidget import WidgetPicker
from GUI.consoleWindow import Console
from GUI.Dialogs.BlueprintsConvertDialog import BlueprintsConvertDialog
from GUI.Dialogs.BlueprintsDialog import BlueprintsDialog
from GUI.Dialogs.BulkFactoryReset import BulkFactoryReset
from GUI.Dialogs.CheckboxMessageBox import CheckboxMessageBox
from GUI.Dialogs.CommandDialog import CommandDialog
from GUI.Dialogs.ConfirmTextDialog import ConfirmTextDialog
from GUI.Dialogs.GeofenceDialog import GeofenceDialog
from GUI.Dialogs.groupManagement import GroupManagement
from GUI.Dialogs.InstalledDevicesDlg import InstalledDevicesDlg
from GUI.Dialogs.LargeTextEntryDialog import LargeTextEntryDialog
from GUI.Dialogs.MultiSelectSearchDlg import MultiSelectSearchDlg
from GUI.Dialogs.NewEndpointDialog import NewEndpointDialog
from GUI.Dialogs.PreferencesDialog import PreferencesDialog
from GUI.Dialogs.ScheduleCmdDialog import ScheduleCmdDialog
from GUI.Dialogs.TemplateDialog import TemplateDialog
from GUI.gridPanel import GridPanel
from GUI.menuBar import ToolMenuBar
from GUI.sidePanel import SidePanel
from GUI.toolBar import ToolsToolBar

from Utility.API.AppUtilities import (
    getAllInstallableApps,
    getAppDictEntry,
    getInstallDevices,
    installAppOnDevices,
    installAppOnGroups,
    uninstallAppOnDevice,
    uninstallAppOnGroup,
)
from Utility.API.BlueprintUtility import (
    checkBlueprintEnabled,
    getAllBlueprints,
    modifyAppsInBlueprints,
    prepareBlueprintClone,
    prepareBlueprintConversion,
    pushBlueprintUpdate,
)
from Utility.API.CommandUtility import createCommand
from Utility.API.DeviceUtility import getAllDevices
from Utility.API.EsperAPICalls import (
    clearAppData,
    getTokenInfo,
    setAppState,
    setKiosk,
    setMulti,
    validateConfiguration,
)
from Utility.API.GroupUtility import getAllGroups, moveGroup
from Utility.API.UserUtility import getAllPendingUsers, getAllUsers, getSpecificUser
from Utility.crypto import crypto
from Utility.EastUtility import (
    TakeAction,
    clearKnownGroups,
    fetchInstalledDevices,
    filterDeviceList,
    getAllDeviceInfo,
    removeNonWhitelisted,
    uploadAppToEndpoint,
)
from Utility.GridActionUtility import bulkFactoryReset, iterateThroughGridRows
from Utility.Logging.ApiToolLogging import ApiToolLog
from Utility.Resource import (
    checkEsperInternetConnection,
    checkForInternetAccess,
    checkIfCurrentThreadStopped,
    correctSaveFileName,
    createNewFile,
    determineDoHereorMainThread,
    displayMessageBox,
    displaySaveDialog,
    getStrRatioSimilarity,
    joinThreadList,
    openWebLinkInBrowser,
    postEventToFrame,
    processFunc,
    resourcePath,
    splitListIntoChunks,
    updateErrorTracker,
)


class NewFrameLayout(wx.Frame):
    def __init__(self):
        self.prefPath = ""
        self.authPath = ""

        self.consoleWin = None
        self.refresh = None
        self.preferences = None
        self.auth_data = None
        self.sleepInhibitor = SleepInhibitor()

        self.WINDOWS = True
        self.isBusy = False
        self.isRunning = False
        self.isSavingPrefs = False
        self.isRunningUpdate = False
        self.isUploading = False
        self.kill = False
        self.CSVUploaded = False
        self.defaultDir = os.getcwd()
        self.gridArrowState = {"next": False, "prev": False}
        self.groups = None
        self.groupManage = None
        self.AppState = None
        self.searchThreads = []
        self.blueprintsEnabled = False
        self.previousGroupFetchThread = None
        self.firstRun = True
        self.changedBlueprints = []

        self.scheduleReport = None
        self.delayScheduleReport = True
        self.scheduleReportRunning = False
        self.scheduleCallLater = []

        self.isSaving = False

        basePath = getToolDataPath()
        if platform.system() == "Windows":
            self.WINDOWS = True
        else:
            self.WINDOWS = False
        self.prefPath = "%s/prefs.json" % basePath
        self.authPath = "%s/auth.csv" % basePath
        self.keyPath = "%s/east.key" % basePath

        wx.Frame.__init__(self, None, title=Globals.TITLE, style=wx.DEFAULT_FRAME_STYLE)
        self.SetSize(Globals.MIN_SIZE)
        self.SetMinSize(Globals.MIN_SIZE)

        self.panel_1 = wx.Panel(self, wx.ID_ANY)

        sizer_4 = wx.FlexGridSizer(1, 2, 0, 0)
        self.sidePanel = SidePanel(self, self.panel_1)
        sizer_4.Add(self.sidePanel, 1, wx.EXPAND, 0)

        self.gridPanel = GridPanel(self, self.panel_1, wx.ID_ANY)
        sizer_4.Add(self.gridPanel, 1, wx.TOP | wx.BOTTOM | wx.RIGHT | wx.EXPAND, 4)

        sizer_4.AddGrowableRow(0)
        sizer_4.AddGrowableCol(1)

        self.panel_1.SetSizer(sizer_4)

        self.Bind(wx.EVT_CLOSE, self.OnQuit)
        self.Bind(wx.EVT_QUERY_END_SESSION, self.OnQuit)
        self.Bind(wx.EVT_END_SESSION, self.OnQuit)
        self.Bind(wx.EVT_END_PROCESS, self.OnQuit)
        self.Bind(wx.EVT_BUTTON, self.onRun, self.sidePanel.runBtn)

        # Menu Bar
        self.menubar = ToolMenuBar(self)
        self.SetMenuBar(self.menubar)
        # Menu Bar end

        # Tool Bar
        self.frame_toolbar = ToolsToolBar(self, -1)
        self.SetToolBar(self.frame_toolbar)
        # Tool Bar end

        # Status Bar
        self.statusBar = ESB.EnhancedStatusBar(self, wx.ID_ANY)
        self.statusBar.SetFieldsCount(2)
        self.SetStatusBar(self.statusBar)
        self.statusBar.SetStatusText("")
        self.statusBar.AddStaicTextAndGauge()
        # End Status Bar

        # Set Icon
        icon = wx.Icon()
        icon.CopyFromBitmap(
            wx.Bitmap(resourcePath("Images/icon.png"), wx.BITMAP_TYPE_PNG)
        )
        self.SetIcon(icon)

        self.notification = None

        self.audit = AuditPosting()

        # Bound Events
        self.DragAcceptFiles(True)
        self.Bind(wx.EVT_DROP_FILES, self.onFileDrop)
        self.Bind(eventUtil.EVT_FETCH, self.onFetch)
        self.Bind(eventUtil.EVT_GROUP, self.addGroupsToGroupChoice)
        self.Bind(eventUtil.EVT_APPS, self.addAppstoAppChoiceThread)
        self.Bind(eventUtil.EVT_COMPLETE, self.onComplete)
        self.Bind(eventUtil.EVT_LOG, self.onLog)
        self.Bind(eventUtil.EVT_COMMAND, self.onCommandDone)
        self.Bind(eventUtil.EVT_UPDATE_GAUGE, self.statusBar.setGaugeValue)
        self.Bind(eventUtil.EVT_UPDATE_TAG_CELL, self.gridPanel.updateTagCell)
        self.Bind(eventUtil.EVT_UPDATE_GRID_CONTENT, self.gridPanel.updateGridContent)
        self.Bind(eventUtil.EVT_UNCHECK_CONSOLE, self.menubar.uncheckConsole)
        self.Bind(eventUtil.EVT_ON_FAILED, self.onFail)
        self.Bind(eventUtil.EVT_CONFIRM_CLONE, self.confirmClone)
        self.Bind(eventUtil.EVT_CONFIRM_CLONE_UPDATE, self.confirmCloneUpdate)
        self.Bind(eventUtil.EVT_MESSAGE_BOX, displayMessageBox)
        self.Bind(eventUtil.EVT_THREAD_WAIT, self.waitForThreadsThenSetCursorDefault)
        self.Bind(eventUtil.EVT_PROCESS_FUNCTION, processFunc)
        self.Bind(eventUtil.EVT_AUDIT, self.audit.postOperation)
        self.Bind(wx.EVT_ACTIVATE_APP, self.MacReopenApp)
        self.Bind(wx.EVT_ACTIVATE, self.onActivate)
        self.Bind(eventUtil.EVT_UPDATE_GAUGE_LATER, self.callSetGaugeLater)
        self.Bind(eventUtil.EVT_DISPLAY_NOTIFICATION, self.displayNotificationEvent)
        self.Bind(wx.EVT_POWER_SUSPENDING, self.onSuspend)

        if self.kill:
            return

        self.prefDialog = PreferencesDialog(parent=self)

        Globals.frame = self

        self.loadPref()
        self.__set_properties()
        self.Layout()
        self.Centre()
        self.tryToMakeActive()

        if self.kill:
            return

        self.menubar.checkCollectionEnabled()
        self.internetCheck = wxThread.GUIThread(
            self, checkForInternetAccess, (self), name="InternetCheck"
        )
        self.internetCheck.startWithRetry()
        self.errorTracker = wxThread.GUIThread(
            self, updateErrorTracker, None, name="updateErrorTracker"
        )
        self.errorTracker.startWithRetry()
        self.menubar.onUpdateCheck(showDlg=True)

        # Display disclaimer unless they have opt'd out.
        if Globals.SHOW_DISCLAIMER:
            self.preferences["showDisclaimer"] = self.menubar.onDisclaimer(
                showCheckBox=True
            )
            Globals.SHOW_DISCLAIMER = self.preferences["showDisclaimer"]

    @api_tool_decorator()
    def tryToMakeActive(self):
        self.Raise()
        self.Iconize(False)
        style = self.GetWindowStyle()
        self.SetWindowStyle(style | wx.STAY_ON_TOP)
        self.SetFocus()
        self.SetWindowStyle(style)

    @api_tool_decorator()
    def __set_properties(self):
        self.SetTitle(Globals.TITLE)
        self.SetBackgroundColour(Color.lightGrey.value)
        self.SetThemeEnabled(False)

        if self.kill:
            return

        maxInt = sys.maxsize
        while True:
            try:
                csv.field_size_limit(maxInt)
                break
            except OverflowError:
                maxInt = int(maxInt / 10)

        self.frame_toolbar.Realize()

    @api_tool_decorator()
    def setFontSizeForLabels(self, parent=None):
        children = None
        if not parent:
            children = self.GetChildren()
        else:
            children = parent.GetChildren()
        for child in children:
            if type(child) == wx.StaticText:
                currentFont = child.GetFont()
                if (
                    currentFont.GetWeight() >= wx.FONTWEIGHT_BOLD
                    and child.Id == wx.ID_BOLD
                ):
                    boldNormalFontSize = ["GridNotebook", "NormalBold"]
                    if child.GetFaceName() in boldNormalFontSize:
                        child.SetFont(
                            wx.Font(
                                Globals.FONT_SIZE,
                                currentFont.GetFamily(),
                                currentFont.GetStyle(),
                                currentFont.GetWeight(),
                                0,
                                "GridNotebook",
                            )
                        )
                    else:
                        child.SetFont(
                            wx.Font(
                                Globals.HEADER_FONT_SIZE,
                                currentFont.GetFamily(),
                                currentFont.GetStyle(),
                                currentFont.GetWeight(),
                                0,
                                "Header",
                            )
                        )
                else:
                    child.SetFont(
                        wx.Font(
                            Globals.FONT_SIZE,
                            currentFont.GetFamily(),
                            currentFont.GetStyle(),
                            currentFont.GetWeight(),
                            0,
                            "Normal",
                        )
                    )
            if child.GetChildren():
                self.setFontSizeForLabels(parent=child)
        determineDoHereorMainThread(self.Refresh)

    @api_tool_decorator()
    def onLog(self, event):
        """ Event trying to log data """
        evtValue = event.GetValue()
        if type(evtValue) is tuple:
            self.Logging(evtValue[0], evtValue[1])
        else:
            self.Logging(evtValue)

    @api_tool_decorator()
    def Logging(self, entry, isError=False):
        """ Frame UI Logging """
        try:
            entry = entry.replace("\n", " ")
            shortMsg = entry
            Globals.LOGLIST.append(entry)
            while len(Globals.LOGLIST) > Globals.MAX_LOG_LIST_SIZE:
                Globals.LOGLIST.pop(0)
            if self.consoleWin:
                self.consoleWin.Logging(entry)
            if "error" in entry.lower():
                isError = True
            if len(entry) >= Globals.MAX_STATUS_CHAR:
                longEntryMsg = "....(See console for details)"
                shortMsg = entry[0 : Globals.MAX_STATUS_CHAR - len(longEntryMsg)]
                shortMsg += longEntryMsg
            self.setStatus(shortMsg, entry, isError)
        except:
            pass

    @api_tool_decorator()
    def AddEndpoint(self, event):
        """ Try to open and load an Auth CSV """
        isValid = False
        errorMsg = None
        name = host = entId = key = None
        while not isValid:
            with NewEndpointDialog(
                errorMsg=errorMsg, name=name, host=host, entId=entId, key=key
            ) as dialog:
                Globals.OPEN_DIALOGS.append(dialog)
                res = dialog.ShowModal()
                Globals.OPEN_DIALOGS.remove(dialog)
                if res == wx.ID_ADD:
                    try:
                        name, host, entId, key, prefix = dialog.getInputValues()
                        csvRow = dialog.getCSVRowEntry()
                        isValid = self.addEndpointEntry(
                            name, host, entId, key, prefix, csvRow
                        )
                        if isValid and type(isValid) == wx.MenuItem:
                            self.loadConfiguartion(isValid)
                    except:
                        displayMessageBox(
                            (
                                "ERROR: An error occured when attempting to add the tenant. Check inputs values and your internet connection.",
                                wx.ICON_ERROR,
                            )
                        )
                else:
                    self.readAuthCSV()
                    if self.auth_data:
                        isValid = True
                    elif res == wx.ID_CANCEL and not self.IsShown():
                        self.OnQuit(event)
                    elif  res == wx.ID_CANCEL:
                        break
        if event and hasattr(event, "Skip"):
            event.Skip()

    def addEndpointEntry(self, name, host, entId, key, prefix, csvRow):
        isValid = validateConfiguration(host, entId, key, prefix=prefix)
        if isValid:
            matchingConfig = []
            if self.auth_data:
                matchingConfig = list(
                    filter(
                        lambda x: x["enterprise"] == entId or x["name"] == name,
                        self.auth_data,
                    )
                )
                if matchingConfig:
                    matchingConfig = matchingConfig[0]
            if (
                not self.auth_data or csvRow not in self.auth_data
            ) and not matchingConfig:
                write_data_to_csv(self.authPath, csvRow, "a")
                Globals.csv_auth_path = self.authPath
                self.readAuthCSV()
                isValid = self.PopulateConfig(auth=self.authPath, getItemForName=name)
                displayMessageBox(("Tenant has been added", wx.ICON_INFORMATION))
            elif csvRow in self.auth_data or matchingConfig:
                self.auth_data_tmp = []
                for entry in self.auth_data:
                    if entry["apiHost"] == matchingConfig["apiHost"]:
                        self.auth_data_tmp.append(csvRow)
                    else:
                        self.auth_data_tmp.append(entry)
                self.auth_data = self.auth_data_tmp

                tmp = [["name", "apiHost", "enterprise", "apiKey", "apiPrefix"]]
                for auth in self.auth_data:
                    authEntry = []
                    indx = 0
                    for val in auth.values():
                        if indx > len(tmp[0]):
                            break
                        authEntry.append(val)
                        indx += 1
                    if authEntry not in tmp:
                        tmp.append(authEntry)
                write_data_to_csv(self.authPath, tmp)
                self.readAuthCSV()
                isValid = self.PopulateConfig(auth=self.authPath, getItemForName=name)
                displayMessageBox(("Tenant has been added", wx.ICON_INFORMATION))
            else:
                displayMessageBox(
                    (
                        "ERROR: Invalid input in Configuration. Check inputs!",
                        wx.ICON_ERROR,
                    )
                )
        return isValid

    @api_tool_decorator()
    def OnQuit(self, e):
        """ Actions to take when frame is closed """
        self.kill = True
        if os.path.exists(self.authPath):
            if self.key and crypto().isFileDecrypt(self.authPath, self.key):
                crypto().encryptFile(self.authPath, self.key)
        if self.consoleWin:
            self.consoleWin.Close()
            self.consoleWin.DestroyLater()
            self.consoleWin = None
        if self.prefDialog:
            self.prefDialog.Close()
            self.prefDialog.DestroyLater()
        if self.notification:
            self.notification.Close()
        if self.menubar.uc:
            self.menubar.uc.Close()
            self.menubar.uc.DestroyLater()
        if e:
            if e.EventType != wx.EVT_CLOSE.typeId:
                self.Close()
        if self.groupManage:
            self.groupManage.Close()
            self.groupManage.DestroyLater()
        thread = ApiToolLog().LogApiRequestOccurrence(
            None, ApiTracker.API_REQUEST_TRACKER, True
        )
        self.savePrefs(self.prefDialog)
        if thread:
            thread.join()
        if hasattr(self, "internetCheck") and self.internetCheck:
            self.internetCheck.stop()
        if hasattr(self, "errorTracker") and self.errorTracker:
            self.errorTracker.stop()
        self.Destroy()
        for item in list(wx.GetTopLevelWindows()):
            if not isinstance(item, NewFrameLayout):
                if item and isinstance(item, wx.Dialog):
                    item.Destroy()
                item.Close()
        Globals.THREAD_POOL.abort()
        wx.Exit()

    @api_tool_decorator()
    def onSaveBoth(self, event):
        self.isSaving = True
        inFile = displaySaveDialog(
            "Save Reports as...",
            "Microsoft Excel Open XML Spreadsheet (*.xlsx)|*.xlsx|CSV files (*.csv)|*.csv"
        )

        if inFile:  # Save button was pressed
            self.setCursorBusy()
            self.toggleEnabledState(False)
            self.gridPanel.disableGridProperties()
            Globals.THREAD_POOL.enqueue(self.saveFile, inFile)
            return True
        else:  # Either the cancel button was pressed or the window was closed
            self.isSaving = False
            return False

    @api_tool_decorator()
    def onSaveBothAll(self, event, action=None):
        if self.sidePanel.selectedDevicesList or self.sidePanel.selectedGroupsList:
            self.isSaving = True
            inFile = displaySaveDialog(
                "Save Reports as...", 
                "Microsoft Excel Open XML Spreadsheet (*.xlsx)|*.xlsx|CSV files (*.csv)|*.csv"
            )

            if inFile:  # Save button was pressed
                self.setCursorBusy()
                self.toggleEnabledState(False)
                self.gridPanel.disableGridProperties()
                self.Logging("Attempting to save file at %s" % inFile)
                self.statusBar.gauge.Pulse()
                Globals.THREAD_POOL.enqueue(self.saveAllFile, inFile, action=action)
                return True
            else:  # Either the cancel button was pressed or the window was closed
                self.isSaving = False
                self.setCursorDefault()
                self.toggleEnabledState(True)
                if self.isRunning:
                    self.isRunning = False
                return False
        else:
            displayMessageBox(
                ("Please select a group and or device(s) first!", wx.OK | wx.ICON_ERROR)
            )

    @api_tool_decorator()
    def getCSVHeaders(self, visibleOnly=False):
        headers = []
        deviceHeaders = Globals.CSV_TAG_ATTR_NAME.keys()
        networkHeaders = Globals.CSV_NETWORK_ATTR_NAME.keys()
        headers.extend(deviceHeaders)
        headers.extend(networkHeaders)
        headersNoDup = []
        [headersNoDup.append(x) for x in headers if x not in headersNoDup]
        headers = headersNoDup

        headersNoDup = []
        for header in headers:
            if visibleOnly:
                if (
                    header in deviceHeaders
                    and self.gridPanel.grid_1.GetColSize(
                        list(deviceHeaders).index(header)
                    )
                    > 0
                    and header not in headersNoDup
                ):
                    headersNoDup.append(header)
                if (
                    header in networkHeaders
                    and self.gridPanel.grid_2.GetColSize(
                        list(networkHeaders).index(header)
                    )
                    > 0
                    and header not in headersNoDup
                ):
                    headersNoDup.append(header)
            else:
                if header in deviceHeaders and header not in headersNoDup:
                    headersNoDup.append(header)
                if header in networkHeaders and header not in headersNoDup:
                    headersNoDup.append(header)
        headers = headersNoDup
        return headers, deviceHeaders, networkHeaders

    @api_tool_decorator()
    def saveAllFile(
        self, inFile, action=None, showDlg=True, allDevices=False, tolarance=1
    ):
        self.sleepInhibitor.inhibit()
        self.start_time = time.time()
        headers, deviceHeaders, networkHeaders = self.getCSVHeaders(
            visibleOnly=Globals.SAVE_VISIBILITY
        )
        self.Logging("Obtaining Device data....")
        deviceList = getAllDeviceInfo(
            self, action=action, allDevices=allDevices, tolarance=tolarance
        )
        gridDeviceData = []
        num = 1
        self.Logging("Processing device information for file")
        for item in deviceList.values():
            if len(item) > 1 and item[1]:
                gridDeviceData.append(item[1])
                self.gridPanel.grid_3_contents += (
                    item[1]["AppsEntry"] if "AppsEntry" in item[1] else []
                )
            postEventToFrame(
                eventUtil.myEVT_UPDATE_GAUGE,
                (int(num / len(deviceList.values())) * 35) + 50,
            )
            num += 1
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 50)

        if hasattr(Globals.frame, "start_time"):
            print(
                "Fetch deviceinfo list time: %s"
                % (time.time() - Globals.frame.start_time)
            )

        self.Logging("Finished compiling information")
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 85)

        self.saveGridData(
            inFile,
            headers,
            deviceHeaders,
            networkHeaders,
            gridDeviceData,
            action=action,
            showDlg=showDlg,
            tolarance=tolarance,
        )
        self.sleepInhibitor.uninhibit()
        postEventToFrame(eventUtil.myEVT_COMPLETE, (True, -1))
        if hasattr(self, "start_time"):
            print("Execution time: %s" % (time.time() - self.start_time))

    @api_tool_decorator()
    def saveFile(self, inFile):
        self.Logging("Preparing to save data to: %s" % inFile)
        self.defaultDir = Path(inFile).parent
        gridDeviceData = []
        headers, deviceHeaders, networkHeaders = self.getCSVHeaders(
            visibleOnly=Globals.SAVE_VISIBILITY
        )
        self.saveGridData(
            inFile,
            headers,
            deviceHeaders,
            networkHeaders,
            gridDeviceData,
            action=GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value,
        )
        postEventToFrame(eventUtil.myEVT_COMPLETE, (True, -1))

    def mergeDeviceAndNetworkInfo(self, device, gridDeviceData):
        tempDict = {}
        tempDict.update(device)
        for entry in self.gridPanel.grid_2_contents:
            if entry["Device Name"] == device[Globals.CSV_TAG_ATTR_NAME["Esper Name"]]:
                tempDict.update(entry)
                break
        gridDeviceData.append(tempDict)

    @api_tool_decorator()
    def saveGridData(
        self,
        inFile,
        headers,
        deviceHeaders,
        networkHeaders,
        gridDeviceData,
        action=None,
        showDlg=True,
        showAppDlg=True,
        renameAppCsv=True,
        tolarance=1,
    ):
        if inFile.endswith(".csv"):
            threads = []
            num = 1
            for device in self.gridPanel.grid_1_contents:
                self.mergeDeviceAndNetworkInfo(device, gridDeviceData)
                val = (num / (len(gridDeviceData) * 2)) * 100
                if val <= 50:
                    postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, int(val))
                num += 1
            joinThreadList(threads)

            gridData = []
            gridData.append(headers)

            createNewFile(inFile)

            num = len(gridDeviceData)
            if not action or action <= GeneralActions.GENERATE_DEVICE_REPORT.value:
                for deviceData in gridDeviceData:
                    rowValues = []
                    for header in headers:
                        value = ""
                        if header in deviceData:
                            value = deviceData[header]
                        else:
                            if header in deviceHeaders:
                                if Globals.CSV_TAG_ATTR_NAME[header] in deviceData:
                                    value = deviceData[
                                        Globals.CSV_TAG_ATTR_NAME[header]
                                    ]
                            if header in networkHeaders:
                                if Globals.CSV_NETWORK_ATTR_NAME[header] in deviceData:
                                    value = deviceData[
                                        Globals.CSV_NETWORK_ATTR_NAME[header]
                                    ]
                        rowValues.append(value)
                    gridData.append(rowValues)
                    val = (num / (len(gridDeviceData) * 2)) * 100
                    if val <= 95:
                        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, int(val))
                    num += 1
                write_data_to_csv(inFile, gridData)
            if (
                not action
                or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
                or action == GeneralActions.GENERATE_APP_REPORT.value
            ) and renameAppCsv:
                if inFile.endswith(".csv"):
                    inFile = inFile[:-4]
                inFile += "_app-report.csv"

            if (
                not action
                or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
                or action == GeneralActions.GENERATE_APP_REPORT.value
            ):
                self.saveAppInfoAsFile(inFile, showAppDlg=showAppDlg)
        elif inFile.endswith(".xlsx"):
            for attempt in range(Globals.MAX_RETRY):
                try:
                    my_wb = xlsxwriter.Workbook(inFile, {"constant_memory": True})
                    # Save Device & Network information, if selected
                    if (
                        not action
                        or action <= GeneralActions.GENERATE_DEVICE_REPORT.value
                        and Globals.COMBINE_DEVICE_AND_NETWORK_SHEETS
                    ):
                        baseSheetName = "Device & Network"
                        if gridDeviceData:
                            self.populateWorkSheet(
                                my_wb,
                                gridDeviceData,
                                baseSheetName,
                                list(deviceHeaders) + list(networkHeaders)[2:],
                                {
                                    **Globals.CSV_TAG_ATTR_NAME,
                                    **Globals.CSV_NETWORK_ATTR_NAME,
                                },
                            )
                        else:
                            self.populateWorkSheet(
                                my_wb,
                                self.mergeDataSource(
                                    self.gridPanel.grid_1_contents,
                                    self.gridPanel.grid_2_contents,
                                ),
                                baseSheetName,
                                list(deviceHeaders) + list(networkHeaders)[2:],
                                {
                                    **Globals.CSV_TAG_ATTR_NAME,
                                    **Globals.CSV_NETWORK_ATTR_NAME,
                                },
                                maxGauge=95,
                                beginGauge=85,
                            )
                    else:
                        # Save Device Info
                        baseSheetName = "Device"
                        if gridDeviceData and (
                            not action
                            or action <= GeneralActions.GENERATE_DEVICE_REPORT.value
                        ):
                            self.populateWorkSheet(
                                my_wb,
                                gridDeviceData,
                                baseSheetName,
                                list(deviceHeaders),
                                Globals.CSV_TAG_ATTR_NAME,
                                maxGauge=90,
                                beginGauge=85,
                            )
                        else:
                            self.populateWorkSheet(
                                my_wb,
                                self.gridPanel.grid_1_contents,
                                baseSheetName,
                                list(deviceHeaders),
                                Globals.CSV_TAG_ATTR_NAME,
                                maxGauge=90,
                                beginGauge=85,
                            )
                        # Save Network Info
                        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 90)
                        baseSheetName = "Network"
                        if gridDeviceData and (
                            not action
                            or action <= GeneralActions.GENERATE_INFO_REPORT.value
                        ):
                            self.populateWorkSheet(
                                my_wb,
                                gridDeviceData,
                                baseSheetName,
                                list(networkHeaders),
                                Globals.CSV_NETWORK_ATTR_NAME,
                                maxGauge=95,
                                beginGauge=90,
                            )
                        else:
                            self.populateWorkSheet(
                                my_wb,
                                self.gridPanel.grid_2_contents,
                                baseSheetName,
                                list(networkHeaders),
                                Globals.CSV_NETWORK_ATTR_NAME,
                                maxGauge=95,
                                beginGauge=90,
                            )
                    postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 95)
                    # Save App Info
                    if self.gridPanel.grid_3_contents and (
                        not action
                        or (
                            action
                            and action == GeneralActions.GENERATE_APP_REPORT.value
                            or action
                            == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
                        )
                    ):
                        baseSheetName = "Application"
                        self.populateWorkSheet(
                            my_wb,
                            self.gridPanel.grid_3_contents,
                            baseSheetName,
                            Globals.CSV_APP_ATTR_NAME,
                            None,
                            maxGauge=100,
                            beginGauge=95,
                        )
                    my_wb.close()
                except Exception as e:
                    if attempt == Globals.MAX_RETRY - 1:
                        raise e
                    else:
                        ApiToolLog().LogError(e)
                        time.sleep(15)
                        self.Logging("Retrying save: %s" % str(e))
                        continue
                break

        Globals.THREAD_POOL.join(tolerance=tolarance)

        self.Logging("---> Info saved to file: " + inFile)
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
        self.gridPanel.enableGridProperties()

        if showDlg:
            res = displayMessageBox(
                (
                    "Report Saved\n\n File saved at: %s\n\nWould you like to navigate to the file?"
                    % inFile,
                    wx.YES_NO | wx.ICON_INFORMATION,
                )
            )
            if res == wx.YES:
                parentDirectory = Path(inFile).parent.absolute()
                openWebLinkInBrowser(parentDirectory, isfile=True)

    def mergeDataSource(self, deviceList, networkList):
        newData = []
        networkIndx = {}
        indx = 0
        for network in networkList:
            if "network_info" in network:
                network = network["network_info"]
            name = network["Esper Name"]
            networkIndx[name] = indx
            indx += 1
        for device in deviceList:
            name = device["EsperName"]
            if name in networkIndx:
                device.update(networkList[networkIndx[name]])
                newData.append(device)
            else:
                newData.append(device)
        return newData

    @api_tool_decorator()
    def populateWorkSheet(
        self,
        workbook,
        dataSource,
        baseSheetName,
        headers,
        headerKeys,
        maxGauge=100,
        beginGauge=85,
    ):
        loopNum = math.ceil(len(dataSource) / Globals.SHEET_CHUNK_SIZE)
        bold = workbook.add_format({"bold": True})
        bold.set_align("center")
        bold.set_align("vcenter")
        numProcessed = 1
        startTime = datetime.now()
        for num in range(loopNum):
            rowIndx = 0
            sheetName = baseSheetName
            if loopNum != 1:
                sheetName += " Part %s" % (num + 1)
            worksheet = workbook.add_worksheet(sheetName)
            maxColumnWidth = {}
            offset = Globals.SHEET_CHUNK_SIZE * num
            endIndx = len(dataSource)
            if len(dataSource) > (offset + Globals.SHEET_CHUNK_SIZE):
                endIndx = offset + Globals.SHEET_CHUNK_SIZE
            colIndx = 0
            for header in headers:
                worksheet.write(rowIndx, colIndx, header, bold)
                maxColumnWidth[header] = len(header) + 5
                colIndx += 1
            rowIndx = 1
            for entry in dataSource[offset:endIndx]:
                colIndx = 0
                for col in headers:
                    value = ""
                    if type(entry) is list:
                        value = entry[colIndx]
                    elif col in entry:
                        value = entry[col]
                    elif headerKeys and col in headerKeys and headerKeys[col] in entry:
                        value = entry[headerKeys[col]]
                    worksheet.write(rowIndx, colIndx, str(value))
                    if headers[colIndx] not in maxColumnWidth or (
                        type(value) is not bool
                        and value is not None
                        and headers[colIndx] in maxColumnWidth
                        and maxColumnWidth[headers[colIndx]] < len(str(value))
                    ):
                        maxColumnWidth[headers[colIndx]] = len(value) + 5
                    elif type(value) is bool or value is None or not value:
                        maxColumnWidth[headers[colIndx]] = len(headers[colIndx])
                    colIndx += 1
                rowIndx += 1
                # Update Gauge every 2 seonds
                if int((datetime.now() - startTime).total_seconds()) % 5 == 0:
                    percent = (
                        int((numProcessed / len(dataSource)) * (maxGauge - beginGauge))
                        + beginGauge
                    )
                    postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, percent)
                numProcessed += 1
            for num in range(len(headers)):
                worksheet.set_column(num, num, maxColumnWidth[headers[num]])

    @api_tool_decorator()
    def saveAppInfoAsFile(self, inFile, showAppDlg=True):
        if self.gridPanel.grid_3_contents:
            gridData = []
            gridData.append(Globals.CSV_APP_ATTR_NAME)
            createNewFile(inFile)

            num = 1
            for entry in self.gridPanel.grid_3_contents:
                if type(entry) is dict:
                    gridData.append(list(entry.values()))
                elif type(entry) is list:
                    for row in entry:
                        gridData.append(list(row.values()))
                val = (num / len(self.gridPanel.grid_3_contents)) * 100
                if val <= 95:
                    postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, int(val))
                num += 1

            write_data_to_csv(inFile, gridData)

            self.Logging("---> Info saved to csv file - " + inFile)
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
            self.toggleEnabledState(True)
            self.setCursorDefault()
            self.gridPanel.enableGridProperties()

            if showAppDlg:
                displayMessageBox(
                    ("Info saved to csv file - " + inFile, wx.OK | wx.ICON_INFORMATION)
                )

    @api_tool_decorator()
    def onUploadCSV(self, event):
        """ Upload device CSV to the device Grid """
        if not Globals.enterprise_id:
            displayMessageBox(
                ("Please load a configuration first!", wx.OK | wx.ICON_ERROR)
            )
            return

        if self.isRunning:
            return

        self.setCursorBusy()
        self.gridPanel.emptyDeviceGrid()
        self.gridPanel.emptyNetworkGrid()
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
        with wx.FileDialog(
            self,
            "Open Device Spreedsheet File",
            wildcard="Spreadsheet Files (*.csv;*.xlsx)|*.csv;*.xlsx|CSV Files (*.csv)|*.csv|Microsoft Excel Open XML Spreadsheet (*.xlsx)|*.xlsx",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
            defaultDir=str(self.defaultDir),
        ) as fileDialog:
            Globals.OPEN_DIALOGS.append(fileDialog)
            result = fileDialog.ShowModal()
            Globals.OPEN_DIALOGS.remove(fileDialog)
            if result == wx.ID_OK:
                # Proceed loading the file chosen by the user
                csv_auth_path = fileDialog.GetPath()
                self.defaultDir = Path(fileDialog.GetPath()).parent
                self.Logging(
                    "--->Attempting to load device data from %s" % csv_auth_path
                )
                self.toggleEnabledState(False)
                if self.WINDOWS:
                    Globals.THREAD_POOL.enqueue(self.openDeviceSpreadsheet, csv_auth_path)
                else:
                    self.openDeviceSpreadsheet(csv_auth_path)
                Globals.THREAD_POOL.enqueue(
                    self.waitForThreadsThenSetCursorDefault,
                    Globals.THREAD_POOL.threads,
                    2,
                    tolerance=1,
                )
            elif result == wx.ID_CANCEL:
                self.setCursorDefault()
                return  # the user changed their mind

    def openDeviceSpreadsheet(self, csv_auth_path):
        self.isUploading = True
        self.Logging("Reading Spreadsheet file: %s" % csv_auth_path)
        if csv_auth_path.endswith(".csv"):
            data = read_data_from_csv(csv_auth_path)
            self.processDeviceCSVUpload(data)
        elif csv_auth_path.endswith(".xlsx"):
            try:
                dfs = self.read_excel_via_openpyxl(csv_auth_path)
                self.processXlsxUpload(dfs)
            except Exception as e:
                print(e)
                pass
        self.gridPanel.notebook_2.SetSelection(0)
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE_LATER, (3000, 0))
        postEventToFrame(
            eventUtil.myEVT_DISPLAY_NOTIFICATION,
            ("E.A.S.T.", "Device CSV Upload Completed"),
        )
        self.setCursorDefault()
        self.toggleEnabledState(True)
        self.sidePanel.groupChoice.Enable(True)
        self.sidePanel.deviceChoice.Enable(True)
        self.gridPanel.enableGridProperties()
        self.gridPanel.thawGridsIfFrozen()
        self.isUploading = False

    def read_excel_via_openpyxl(self, path: str) -> pd.DataFrame:
        # Load the Excel file
        # start = time.time()
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # print("openpyxl load time: %s" % (time.time() - start))
        df = None
        rows = []
        for sheet in workbook.sheetnames:
            if "Device & Network" in sheet or "Device and Network" in sheet or "Device" in sheet:
                # Select the worksheet
                worksheet = workbook[sheet]
                # Extract the data
                for row in worksheet.iter_rows(values_only=True):
                    rows.append(row)
        # Convert to pandas dataframe
        df = pd.DataFrame(rows[1:], columns=rows[0])
        # print("openpyxl process time: %s" % (time.time() - start))
        return df

    def processXlsxUpload(self, data):
        self.CSVUploaded = True
        self.toggleEnabledState(False)
        self.sidePanel.groupChoice.Enable(False)
        self.sidePanel.deviceChoice.Enable(False)
        self.gridPanel.disableGridProperties()
        self.gridPanel.freezeGrids()
        dataList = data.values.tolist()
        dataList.insert(0, list(data.keys()))
        self.Logging("Processing Spreadsheet data...")
        self.processSpreadsheetDataByGrid(
            self.gridPanel.grid_1,
            dataList,
            Globals.CSV_TAG_ATTR_NAME,
            Globals.grid1_lock,
        )

    def processDeviceCSVUpload(self, data):
        self.CSVUploaded = True
        self.toggleEnabledState(False)
        self.sidePanel.groupChoice.Enable(False)
        self.sidePanel.deviceChoice.Enable(False)
        self.gridPanel.disableGridProperties()
        self.gridPanel.freezeGrids()
        self.Logging("Processing Spreadsheet data...")
        self.processSpreadsheetDataByGrid(
            self.gridPanel.grid_1,
            data,
            Globals.CSV_TAG_ATTR_NAME,
            Globals.grid1_lock,
        )

    @api_tool_decorator(locks=[Globals.grid1_lock, Globals.grid2_lock])
    def processSpreadsheetDataByGrid(self, grid, data, headers, lock=None):
        if lock:
            lock.acquire()
        grid_headers = list(headers.keys())
        len_reader = len(data)
        rowCount = 1
        header = data[0]
        validHeader = False
        for colName in header:
            if colName in grid_headers:
                # Assume its valid if it has at least one valid entry
                validHeader = True
                break
        if not validHeader:
            raise Exception(
                "Not Able to Process Spreadsheet File! Missing HEADERS!"
            )
        grid.AppendRows(len(data) - 1)
        rowNum = 0
        for row in data[1:]:
            postEventToFrame(
                eventUtil.myEVT_UPDATE_GAUGE, int(float((rowCount) / len_reader) * 100)
            )
            rowCount += 1
            if not all("" == val or val.isspace() for val in row):
                self.processSpreadsheetRow(row, rowNum, header, grid_headers, grid)
                rowNum += 1
            else:
                grid.DeleteRows(0, grid.GetNumberRows() - 1)
        if lock:
            lock.release()

    def processSpreadsheetRow(self, row, rowNum, header, grid_headers, grid):
        fileCol = 0
        for colValue in row:
            # Get Column name from header
            colName = self.getAndValidateColumnHeaderName(header, fileCol)
            if colValue == "--":
                colValue = ""
            # Skip column if 1) Column name is not in the grid headers 2) Column name is in the deprecated headers
            if (
                fileCol < len(header)
                and colName.strip() in Globals.CSV_DEPRECATED_HEADER_LABEL
            ) or (
                fileCol < len(header)
                and colName.strip() not in grid_headers
                and colName != "devicename"
            ):
                fileCol += 1
                continue
            expectedCol = ""
            # Check to see if Column Name is exact match
            if colName in grid_headers:
                expectedCol = colName
            else:
                # If not try to find close match
                for col in grid_headers:
                    if getStrRatioSimilarity(colName, col) >= 95:
                        expectedCol = col
                        break
            # Proceed if we have a valid Column name
            if expectedCol:
                toolCol = grid_headers.index(expectedCol)
                # Attempt to process Tag column value into proper value
                if expectedCol == "Tags":
                    try:
                        ast.literal_eval(colValue)
                    except:
                        colValue = str(colValue)
                        if (
                            expectedCol == "Tags"
                            and "," in colValue
                            and (
                                (
                                    colValue.count('"') % 2 != 0
                                    or colValue.count("'") % 2 != 0
                                    or colValue.count("’") % 2 != 0
                                )
                                or (
                                    '"' not in colValue
                                    and "'" not in colValue
                                    and "’" not in colValue
                                )
                            )
                        ):
                            colValue = '"' + colValue + '"'
                isEditable = True
                if grid_headers[toolCol] in Globals.CSV_EDITABLE_COL:
                    isEditable = False
                addColToGridRow(grid, rowNum, toolCol, str(colValue), isEditable)
                fileCol += 1

    def getAndValidateColumnHeaderName(self, header, indx):
        colName = str(header[indx]) if len(header) > indx else ""
        if colName.lower() == "storenumber" or colName.lower() == "store number":
            colName = "Alias"
        if colName.lower() == "tag":
            colName = "Tags"
        return colName

    @api_tool_decorator()
    def PopulateConfig(self, auth=None, getItemForName=None):
        """Populates Configuration From CSV"""
        self.Logging("--->Loading Tenants from %s" % Globals.csv_auth_path)
        if auth:
            if Globals.csv_auth_path != auth:
                Globals.csv_auth_path = auth
        configfile = Globals.csv_auth_path

        for item in self.menubar.configMenuOptions:
            try:
                self.menubar.configMenu.Delete(item)
            except:
                pass
        self.menubar.configMenuOptions = []

        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
        returnItem = None
        if os.path.isfile(configfile):
            if self.auth_data:
                auth_csv_reader = self.auth_data
                maxRow = len(auth_csv_reader)
                num = 1

                # Handle empty File
                if maxRow == 0:
                    self.Logging("--->ERROR: Empty Auth File, please add an Tenant!")
                    self.AddEndpoint(None)
                    return

                for row in auth_csv_reader:
                    postEventToFrame(
                        eventUtil.myEVT_UPDATE_GAUGE, int(float(num / maxRow) * 25)
                    )
                    num += 1
                    if "name" in row:
                        self.sidePanel.configChoice[row["name"]] = row
                        item = self.menubar.configMenu.Append(
                            wx.ID_ANY, row["name"], row["name"], kind=wx.ITEM_CHECK
                        )
                        self.Bind(wx.EVT_MENU, self.loadConfiguartion, item)
                        self.menubar.configMenuOptions.append(item)
                        if (
                            str(getItemForName) == row["name"]
                            or str(getItemForName).lower() == row["name"].lower()
                        ):
                            returnItem = item
                    else:
                        self.Logging(
                            "--->ERROR: Please check that the Auth CSV is set up correctly!"
                        )
                        defaultConfigVal = self.menubar.configMenu.Append(
                            wx.ID_NONE,
                            "No Loaded Tenants",
                            "No Loaded Tenants",
                        )
                        self.menubar.configMenuOptions.append(defaultConfigVal)
                        self.Bind(wx.EVT_MENU, self.AddEndpoint, defaultConfigVal)
                        return
            self.Logging(
                "---> Please Select an Tenant From the Configuartion Menu (defaulting to first Config)"
            )
            indx = 0
            if type(Globals.LAST_OPENED_ENDPOINT) is str:
                found = False
                for item in self.menubar.configMenuOptions:
                    if item.GetItemLabelText() == Globals.LAST_OPENED_ENDPOINT:
                        found = True
                        break
                    indx += 1
                if not found:
                    indx = 0
            else:
                indx = Globals.LAST_OPENED_ENDPOINT
                if indx > len(self.menubar.configMenuOptions):
                    indx = len(self.menubar.configMenuOptions) - 1
                    Globals.LAST_OPENED_ENDPOINT = indx
            if indx >= 0 and len(self.menubar.configMenuOptions) > indx:
                defaultConfigItem = self.menubar.configMenuOptions[indx]
                defaultConfigItem.Check(True)
                self.loadConfiguartion(defaultConfigItem)
        else:
            self.Logging(
                "---> "
                + configfile
                + " not found - PLEASE Quit and create configuration file"
            )
            defaultConfigVal = self.menubar.configMenu.Append(
                wx.ID_NONE, "No Loaded Tenants", "No Loaded Tenants"
            )
            self.menubar.configMenuOptions.append(defaultConfigVal)
            self.Bind(wx.EVT_MENU, self.AddEndpoint, defaultConfigVal)
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            (wx.CallLater, (300, self.statusBar.setGaugeValue, 0)),
        )
        return returnItem

    @api_tool_decorator()
    def setCursorDefault(self):
        """ Set cursor icon to default state """
        try:
            self.isBusy = False
            myCursor = wx.Cursor(wx.CURSOR_DEFAULT)
            self.SetCursor(myCursor)
        except:
            pass

    @api_tool_decorator()
    def setCursorBusy(self):
        """ Set cursor icon to busy state """
        self.isBusy = True
        myCursor = wx.Cursor(wx.CURSOR_WAIT)
        self.SetCursor(myCursor)

    @api_tool_decorator()
    def loadConfiguartion(self, event, *args, **kwargs):
        """Populate Frame Layout With Device Configuration"""
        self.configMenuItem = self.menubar.configMenu.FindItemById(event.Id)
        if not self.firstRun:
            Globals.THREAD_POOL.clearQueue()
        self.onClearGrids()
        clearKnownGroups()
        try:
            if self.groupManage:
                self.groupManage.Destroy()
        except:
            pass
        # Reset Side Panel
        self.sidePanel.groups = {}
        self.sidePanel.devices = {}
        self.sidePanel.groupDeviceCount = {}
        self.sidePanel.clearSelections(clearApp=True)
        self.sidePanel.destroyMultiChoiceDialogs()
        self.sidePanel.deviceChoice.Enable(False)
        self.sidePanel.removeEndpointBtn.Enable(False)
        self.sidePanel.notebook_1.SetSelection(0)
        self.sidePanel.clearStoredApps()
        # Clear Search Input
        self.frame_toolbar.search.SetValue("")
        # Disable other options
        self.toggleEnabledState(False)
        self.setCursorBusy()

        self.firstRun = False

        if ApiTracker.API_REQUEST_SESSION_TRACKER > 0:
            thread = ApiToolLog().LogApiRequestOccurrence(
                None, ApiTracker.API_REQUEST_TRACKER, True
            )
            if thread:
                thread.join()
            resetDict = {}
            for key in ApiTracker.API_REQUEST_TRACKER.keys():
                resetDict[key] = 0
            ApiTracker.API_REQUEST_TRACKER = resetDict
            ApiTracker.API_REQUEST_SESSION_TRACKER = 0
        try:
            self.Logging(
                "--->Attempting to load configuration: %s."
                % self.configMenuItem.GetItemLabelText()
            )
            selectedConfig = self.sidePanel.configChoice[
                self.configMenuItem.GetItemLabelText()
            ]

            indx = 0
            found = False
            foundItem = None
            for item in self.configMenuItem.Menu.MenuItems:
                if item != self.configMenuItem:
                    item.Check(False)
                else:
                    item.Check(True)
                    found = True
                    foundItem = item
                if not found:
                    indx += 1
            Globals.LAST_OPENED_ENDPOINT = foundItem.GetItemLabelText()
            if self.prefDialog:
                self.prefDialog.SetPref("last_endpoint", Globals.LAST_OPENED_ENDPOINT)
                Globals.THREAD_POOL.enqueue(self.savePrefs, self.prefDialog)

            filledIn = False
            for _ in range(Globals.MAX_RETRY):
                if not filledIn:
                    filledIn = self.fillInConfigListing(selectedConfig)
                else:
                    break

            if not filledIn:
                raise Exception("Failed to load configuration")
        except Exception as e:
            self.Logging(
                "---> An Error has occured while loading the configuration, please try again."
            )
            print(e)
            ApiToolLog().LogError(e)
            self.configMenuItem.Check(False)

    @api_tool_decorator()
    def fillInConfigListing(self, config):
        self.sidePanel.configList.Clear()
        host = key = entId = prefix = ""
        if type(config) is dict or isinstance(config, dict):
            host = config["apiHost"]
            key = config["apiKey"]
            prefix = config["apiPrefix"]
            entId = config["enterprise"]
        elif type(config) is tuple or isinstance(config, tuple):
            host = config[0]
            key = config[1]
            prefix = config[2]
            entId = config[3]

        if not prefix:
            prefix = "Bearer"

        if host and key and prefix and entId:
            self.sidePanel.configList.AppendText("API Host = " + host + "\n")
            self.sidePanel.configList.AppendText("API key = " + key + "\n")
            self.sidePanel.configList.AppendText("API Prefix = " + prefix + "\n")
            self.sidePanel.configList.AppendText("Enterprise = " + entId)
            self.sidePanel.configList.ShowPosition(0)
            Globals.IS_TOKEN_VALID = False

            if "https" in str(host):
                Globals.configuration.host = host.strip()
                Globals.configuration.api_key["Authorization"] = key.strip()
                Globals.configuration.api_key_prefix["Authorization"] = prefix.strip()
                Globals.enterprise_id = entId.strip()

                Globals.THREAD_POOL.enqueue(self.validateToken)

                postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 50)
                self.menubar.toggleCloneMenuOptions(False, True)
                if Globals.HAS_INTERNET is None:
                    Globals.HAS_INTERNET = checkEsperInternetConnection()
                threads = []
                if Globals.HAS_INTERNET:
                    self.toggleEnabledState(False)
                    groupThread = self.PopulateGroups()
                    appThread = self.PopulateApps()
                    blueprints = wxThread.GUIThread(
                        self,
                        self.loadConfigCheckBlueprint,
                        config,
                        name="loadConfigCheckBlueprint",
                    )
                    blueprints.start()
                    threads = [groupThread, appThread, blueprints]
                Globals.THREAD_POOL.enqueue(
                    self.waitForThreadsThenSetCursorDefault,
                    threads,
                    0,
                    tolerance=1,
                )
                return True
        else:
            displayMessageBox(("Invalid Configuration", wx.ICON_ERROR))
            return False

    @api_tool_decorator(locks=[Globals.token_lock])
    def validateToken(self):
        Globals.token_lock.acquire()
        try:
            res = getTokenInfo(maxAttempt=2)
        except:
            pass
        if res and hasattr(res, "expires_on"):
            Globals.IS_TOKEN_VALID = True
            if res.expires_on <= datetime.now(res.expires_on.tzinfo) or not res:
                Globals.IS_TOKEN_VALID = False
                determineDoHereorMainThread(self.promptForNewToken)
        elif (
            res
            and hasattr(res, "body")
            and "Authentication credentials were not provided" in res.body
        ):
            Globals.IS_TOKEN_VALID = False
            determineDoHereorMainThread(self.promptForNewToken)

        if res and hasattr(res, "user"):
            Globals.TOKEN_USER = getSpecificUser(res.user)
            if res and hasattr(res, "scope"):
                if "write" not in res.scope:
                    self.menubar.fileAddUser.Enable(False)
        if Globals.token_lock.locked():
            Globals.token_lock.release()

    def promptForNewToken(self):
        newToken = ""
        while not newToken:
            with TextEntryDialog(
                self,
                "Please enter a new API Token for %s" % Globals.configuration.host,
                "%s - API Token has expired!" % self.configMenuItem.GetItemLabelText(),
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                if dlg.ShowModal() == wx.ID_OK:
                    newToken = dlg.GetValue()
                else:
                    Globals.OPEN_DIALOGS.remove(dlg)
                    break
                Globals.OPEN_DIALOGS.remove(dlg)
            newToken = newToken.strip()
            if newToken:
                csvRow = [
                    self.configMenuItem.GetItemLabelText(),
                    Globals.configuration.host,
                    Globals.enterprise_id,
                    newToken,
                    Globals.configuration.api_key_prefix["Authorization"],
                ]
                valid = self.addEndpointEntry(
                    self.configMenuItem.GetItemLabelText(),
                    Globals.configuration.host,
                    Globals.enterprise_id,
                    newToken,
                    Globals.configuration.api_key_prefix["Authorization"],
                    csvRow,
                )
                if not valid:
                    newToken = ""

    @api_tool_decorator()
    def waitForThreadsThenSetCursorDefault(
        self, threads, source=None, action=None, tolerance=0
    ):
        if hasattr(threads, "GetValue"):
            evtVal = threads.GetValue()
            threads = evtVal[0]
            if len(evtVal) > 1:
                source = evtVal[1]
            if len(evtVal) > 2:
                action = evtVal[2]
        if threads == Globals.THREAD_POOL.threads:
            time.sleep(3)
            Globals.THREAD_POOL.join(tolerance=tolerance)
        else:
            joinThreadList(threads)
        if source == 0:
            self.gridPanel.setColVisibility()
            self.sidePanel.sortAndPopulateAppChoice()
            self.sidePanel.groupChoice.Enable(True)
            self.sidePanel.actionChoice.Enable(True)
            self.sidePanel.removeEndpointBtn.Enable(True)
            self.handleScheduleReportPref()
        if source == 1:
            if not self.sidePanel.devices:
                self.sidePanel.selectedDevices.Append("No Devices Found", "")
                self.sidePanel.deviceChoice.Enable(False)
                self.menubar.setSaveMenuOptionsEnableState(False)
                self.menubar.enableConfigMenu()
                self.Logging("---> No Devices found")
            else:
                self.sidePanel.deviceChoice.Enable(True)
                self.sidePanel.sortAndPopulateAppChoice()
                self.Logging("---> Application list populated")
                if not self.isRunning:
                    self.menubar.enableConfigMenu()
                self.menubar.setSaveMenuOptionsEnableState(True)
            if (
                not self.preferences
                or (
                    "enableDevice" in self.preferences
                    and self.preferences["enableDevice"]
                )
            ) and self.sidePanel.devices:
                self.sidePanel.deviceChoice.Enable(True)
            else:
                self.sidePanel.deviceChoice.Enable(False)
            self.displayNotification("Finished loading devices", "")
        if source == 2:
            indx = self.sidePanel.actionChoice.GetItems().index(
                list(Globals.GRID_ACTIONS.keys())[0]
            )
            self.isUploading = False
            if self.sidePanel.actionChoice.GetSelection() < indx:
                self.sidePanel.actionChoice.SetSelection(indx)
            determineDoHereorMainThread(self.gridPanel.thawGridsIfFrozen)
            determineDoHereorMainThread(self.gridPanel.enableGridProperties)
            determineDoHereorMainThread(self.gridPanel.autoSizeGridsColumns)
            determineDoHereorMainThread(self.sidePanel.groupChoice.Enable, True)
            determineDoHereorMainThread(self.sidePanel.deviceChoice.Enable, True)
        if source == 3:
            cmdResults = []
            if (
                action == GeneralActions.SET_KIOSK.value
                or action == GeneralActions.SET_MULTI.value
                or action == GeneralActions.SET_APP_STATE.value
                or action == GeneralActions.REMOVE_NON_WHITELIST_AP.value
                or action == GeneralActions.MOVE_GROUP.value
                or action == GeneralActions.INSTALL_APP.value
                or action == GeneralActions.UNINSTALL_APP.value
                or action == GridActions.SET_APP_STATE.value
                or action == GridActions.MOVE_GROUP.value
                or action == GridActions.FACTORY_RESET.value
            ):
                if threads == Globals.THREAD_POOL.threads:
                    resp = Globals.THREAD_POOL.results()
                    for t in resp:
                        if t:
                            if type(t) == list:
                                cmdResults = cmdResults + t
                            else:
                                cmdResults.append(t)
                else:
                    for t in threads:
                        if t.result:
                            if type(t.result) == list:
                                cmdResults = cmdResults + t.result
                            else:
                                cmdResults.append(t.result)
            if action and action == GeneralActions.CLEAR_APP_DATA.value:
                if threads == Globals.THREAD_POOL.threads:
                    resp = Globals.THREAD_POOL.results()
                    for t in resp:
                        if t:
                            displayMessageBox(
                                (
                                    "Clear App Data Command has been sent to the device(s).\nPlease check devices' event feeds for command status.",
                                    wx.ICON_INFORMATION,
                                )
                            )
                            break
                else:
                    for t in threads:
                        if t.result:
                            displayMessageBox(
                                (
                                    "Clear App Data Command has been sent to the device(s).\nPlease check devices' event feeds for command status.",
                                    wx.ICON_INFORMATION,
                                )
                            )
                            break
            self.onComplete((True, action, cmdResults))
        self.toggleEnabledState(not self.isRunning and not self.isSavingPrefs)
        self.setCursorDefault()
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            self.Refresh,
        )

    @api_tool_decorator()
    def PopulateGroups(self):
        """ Populate Group Choice """
        self.sidePanel.groupChoice.Enable(False)
        self.Logging("--->Attempting to populate groups...")
        self.setCursorBusy()
        if self.previousGroupFetchThread:
            self.previousGroupFetchThread.stop()
        thread = wxThread.GUIThread(
            self,
            getAllGroups,
            (
                "",
                None,
                None,
                Globals.MAX_RETRY,
                1,
            ),
            eventType=eventUtil.myEVT_GROUP,
            name="PopulateGroupsGetAll",
        )
        thread.startWithRetry()
        self.previousGroupFetchThread = thread
        return thread

    @api_tool_decorator()
    def addGroupsToGroupChoice(self, event):
        """ Populate Group Choice """
        self.sidePanel.groupsResp = event.GetValue()
        results = None
        if hasattr(self.sidePanel.groupsResp, "results"):
            results = self.sidePanel.groupsResp.results
        elif (
            type(self.sidePanel.groupsResp) is dict
            and "results" in self.sidePanel.groupsResp
        ):
            results = self.sidePanel.groupsResp["results"]
        num = 1
        self.groups = results
        if results:
            results = sorted(
                results,
                key=lambda i: i.name.lower()
                if hasattr(i, "name")
                else i["name"].lower()
                if type(i) is dict
                else i,
            )
        if results and len(results):
            for group in results:
                if hasattr(group, "name"):
                    if (
                        hasattr(group, "enterprise")
                        and Globals.enterprise_id not in group.enterprise
                    ):
                        return
                    if group.name not in self.sidePanel.groups:
                        self.sidePanel.groups[group.name] = group.id
                    else:
                        self.sidePanel.groups[group.path] = group.id
                    if group.id not in Globals.knownGroups:
                        Globals.knownGroups[group.id] = group
                elif type(group) is dict:
                    if Globals.enterprise_id not in group["enterprise"]:
                        return
                    groupEntryId = group["name"]
                    if groupEntryId not in self.sidePanel.groups:
                        self.sidePanel.groups[groupEntryId] = group["id"]
                    else:
                        groupEntryId = group["path"]
                        self.sidePanel.groups[groupEntryId] = group["id"]

                    pathParts = group["path"].split("/")
                    path = ""
                    for p in pathParts:
                        if path:
                            path = path + "/" + p
                        else:
                            path = p
                        if path in self.sidePanel.groupDeviceCount:
                            self.sidePanel.groupDeviceCount[path] += group[
                                "device_count"
                            ]
                        else:
                            self.sidePanel.groupDeviceCount[path] = group[
                                "device_count"
                            ]

                    if group["id"] not in Globals.knownGroups:
                        Globals.knownGroups[group["id"]] = group
                postEventToFrame(
                    eventUtil.myEVT_UPDATE_GAUGE,
                    50 + int(float(num / len(results)) * 25),
                )
                num += 1
        self.sidePanel.groupChoice.Enable(True)
        self.sidePanel.actionChoice.Enable(True)
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            self.Refresh,
        )

    @api_tool_decorator()
    def PopulateDevices(self, event):
        """ Populate Device Choice """
        self.menubar.setSaveMenuOptionsEnableState(False)
        self.SetFocus()
        self.Logging("--->Attempting to populate devices of selected group(s)")
        self.sidePanel.deviceChoice.Enable(False)
        self.setCursorBusy()
        if not self.preferences or (
            "enableDevice" in self.preferences and self.preferences["enableDevice"]
        ):
            self.sidePanel.runBtn.Enable(False)
            self.frame_toolbar.EnableTool(self.frame_toolbar.rtool.Id, False)
            self.frame_toolbar.EnableTool(self.frame_toolbar.cmdtool.Id, False)
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
            self.statusBar.gauge.Pulse()
        else:
            if not self.isRunning or not self.isBusy:
                self.sidePanel.runBtn.Enable(True)
                self.frame_toolbar.EnableTool(self.frame_toolbar.rtool.Id, True)
        self.frame_toolbar.EnableTool(self.frame_toolbar.cmdtool.Id, True)
        self.toggleEnabledState(False)
        Globals.THREAD_POOL.enqueue(self.addDevicesToDeviceChoice, tolerance=2)
        Globals.THREAD_POOL.enqueue(
            self.waitForThreadsThenSetCursorDefault,
            Globals.THREAD_POOL.threads,
            1,
            tolerance=1,
        )

    @api_tool_decorator()
    def addDevicesToDeviceChoice(self, tolerance=0):
        """ Populate Device Choice """
        for clientData in self.sidePanel.selectedGroupsList:
            api_response = getAllDevices(
                clientData,
                limit=Globals.limit,
                fetchAll=Globals.GROUP_FETCH_ALL,
                tolarance=tolerance,
            )
            self.sidePanel.deviceResp = api_response
            splitResults = None
            if hasattr(api_response, "results") and len(api_response.results):
                self.Logging("---> Processing fetched devices...")
                if not Globals.SHOW_DISABLED_DEVICES:
                    api_response.results = list(
                        filter(filterDeviceList, api_response.results)
                    )
                api_response.results = sorted(
                    api_response.results,
                    key=lambda i: i.device_name.lower(),
                )
                splitResults = splitListIntoChunks(api_response.results)
            elif type(api_response) is dict and len(api_response["results"]):
                self.Logging("---> Processing fetched devices...")
                if not Globals.SHOW_DISABLED_DEVICES:
                    api_response["results"] = list(
                        filter(filterDeviceList, api_response["results"])
                    )
                api_response["results"] = sorted(
                    api_response["results"],
                    key=lambda i: i["device_name"].lower(),
                )
                splitResults = splitListIntoChunks(api_response["results"])

            if splitResults:
                for chunk in splitResults:
                    Globals.THREAD_POOL.enqueue(self.processAddDeviceToChoice, chunk)
                Globals.THREAD_POOL.join(tolerance=tolerance)

    def processAddDeviceToChoice(self, chunk):
        for device in chunk:
            name = ""
            if hasattr(device, "hardware_info"):
                name = "%s ~ %s ~ %s %s" % (
                    device.hardware_info["manufacturer"],
                    device.hardware_info["model"],
                    device.device_name,
                    "~ %s" % device.alias_name if device.alias_name else "",
                )
            else:
                name = "%s ~ %s ~ %s %s" % (
                    device["hardwareInfo"]["manufacturer"],
                    device["hardwareInfo"]["model"],
                    device["device_name"],
                    "~ %s" % device["alias_name"] if device["alias_name"] else "",
                )
            if name and name not in self.sidePanel.devices:
                if hasattr(device, "id"):
                    self.sidePanel.devices[name] = device.id
                else:
                    self.sidePanel.devices[name] = device["id"]

    @api_tool_decorator()
    def PopulateApps(self):
        """ Populate App Choice """
        self.Logging("--->Attempting to populate apps...")
        self.sidePanel.appChoice.Enable(False)
        self.setCursorBusy()
        thread = wxThread.GUIThread(
            self, self.fetchAllInstallableApps, None, name="PopulateApps"
        )
        thread.startWithRetry()
        return thread

    @api_tool_decorator(locks=[Globals.token_lock])
    def fetchAllInstallableApps(self):
        Globals.token_lock.acquire()
        Globals.token_lock.release()
        if Globals.IS_TOKEN_VALID:
            resp = getAllInstallableApps(tolerance=1)
            self.addAppsToAppChoice(resp)

    def addAppstoAppChoiceThread(self, event):
        api_response = None
        if hasattr(event, "GetValue"):
            api_response = event.GetValue()
        else:
            api_response = event
        if hasattr(api_response, "results"):
            api_response = api_response.results
        elif type(api_response) == tuple and "results" in api_response[1]:
            api_response = api_response[1]["results"]
        elif type(api_response) == dict:
            api_response = api_response["results"]
        if api_response:
            Globals.THREAD_POOL.enqueue(self.addAppsToAppChoice, api_response)

    @api_tool_decorator()
    def addAppsToAppChoice(self, event):
        """ Populate App Choice """
        api_response = None
        if hasattr(event, "GetValue"):
            api_response = event.GetValue()
        else:
            api_response = event
        results = None

        if hasattr(api_response, "results"):
            results = api_response.results
        elif type(api_response) == tuple and "results" in api_response[1]:
            results = api_response[1]["results"]
        elif type(api_response) == dict:
            results = api_response["results"]

        if results and type(results[0]) == dict and "application" in results[0]:
            results = sorted(
                results,
                key=lambda i: i["application"]["application_name"].lower(),
            )
        elif results and hasattr(results[0], "application_name"):
            results = sorted(
                results,
                key=lambda i: i.application_name.lower(),
            )
        elif results and type(results[0]) == dict and "application_name" in results[0]:
            results = sorted(
                results,
                key=lambda i: i["application_name"].lower(),
            )
        elif results:
            results = sorted(
                results,
                key=lambda i: i["app_name"].lower(),
            )

        if results and len(results):
            for app in results:
                entry = getAppDictEntry(app)
                if (
                    entry
                    and entry not in self.sidePanel.enterpriseApps
                    and ("isValid" in entry and entry["isValid"])
                ):
                    self.sidePanel.enterpriseApps.append(entry)

    @api_tool_decorator()
    def getAppDataForRun(self):
        appSelection = self.sidePanel.selectedApp.GetSelection()
        appLabel = (
            self.sidePanel.selectedAppEntry["name"]
            if self.sidePanel.selectedAppEntry
            else ""
        )
        return appSelection, appLabel

    @api_tool_decorator()
    def onRun(self, event=None):
        """ Try to run the specifed Action on a group or device """
        if not Globals.HAS_INTERNET:
            displayMessageBox(
                (
                    "ERROR: An internet connection is required when using the tool!",
                    wx.OK | wx.ICON_ERROR | wx.CENTRE,
                )
            )
            return

        if (
            self.isBusy
            or not self.sidePanel.runBtn.IsEnabled()
            or self.scheduleReportRunning
        ):
            return

        end_time = time.time() + 120
        while (
            self.isRunning
            or self.isRunningUpdate
            or self.isSavingPrefs
            or self.isUploading
            or self.isBusy
        ) and time.time() < end_time:
            time.sleep(1)
        self.start_time = time.time()
        self.setCursorBusy()
        self.isRunning = True
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
        self.toggleEnabledState(False)
        self.AppState = None
        self.sleepInhibitor.inhibit()

        self.gridPanel.grid_1.UnsetSortingColumn()
        self.gridPanel.grid_2.UnsetSortingColumn()
        self.gridPanel.grid_3.UnsetSortingColumn()

        appSelection, appLabel = self.getAppDataForRun()
        actionSelection = self.sidePanel.actionChoice.GetSelection()
        actionClientData = self.sidePanel.actionChoice.GetClientData(actionSelection)

        allDevicesSelected = (
            True
            if self.sidePanel.deviceResp
            and len(self.sidePanel.selectedDevicesList)
            == len(self.sidePanel.deviceResp["results"])
            else False
        )

        actionLabel = (
            self.sidePanel.actionChoice.Items[actionSelection]
            if len(self.sidePanel.actionChoice.Items) > 0
            and self.sidePanel.actionChoice.Items[actionSelection]
            else ""
        )

        estimatedDeviceCount = len(self.sidePanel.selectedDevicesList)
        if not self.sidePanel.selectedDevicesList:
            for group in self.sidePanel.selectedGroupsList:
                match = list(filter(lambda x: x["id"] == group, self.groups))
                if match:
                    match = match[0]
                    if type(match) == dict and "device_count" in match:
                        match = list(
                            filter(
                                lambda x: x.endswith(match["name"]),
                                self.sidePanel.groupDeviceCount.keys(),
                            )
                        )
                        if match:
                            match = match[0]
                            estimatedDeviceCount += self.sidePanel.groupDeviceCount[
                                match
                            ]
                    elif hasattr(match, "device_count"):
                        match = list(
                            filter(
                                lambda x: x.endswith(match.name),
                                self.sidePanel.groupDeviceCount.keys(),
                            )
                        )
                        if match:
                            match = match[0]
                            estimatedDeviceCount += self.sidePanel.groupDeviceCount[
                                match
                            ]

        if (
            actionClientData < GeneralActions.GENERATE_APP_REPORT.value
            and estimatedDeviceCount > Globals.MAX_DEVICE_COUNT
        ) or (
            (actionClientData == GeneralActions.GENERATE_APP_REPORT.value
            or actionClientData == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value)
            and estimatedDeviceCount > (Globals.MAX_DEVICE_COUNT / 25)
        ):
            res = displayMessageBox(
                (
                    "Looks like you are generating a report for a large subset of devices.\n\nWe will save the info directly to a file.",
                    wx.ICON_INFORMATION | wx.CENTRE | wx.OK,
                )
            )
            if res == wx.OK:
                self.onClearGrids()
                self.gridPanel.grid_1_contents = []
                self.gridPanel.grid_2_contents = []
                self.gridPanel.grid_3_contents = []
                self.gridPanel.userEdited = []
                return self.onSaveBothAll(None, action=actionClientData)
            else:
                return

        if actionClientData == GeneralActions.REMOVE_NON_WHITELIST_AP.value:
            with LargeTextEntryDialog(
                self,
                "Enter Wifi SSIDs you want whitelisted, as a comma seperated list:",
                "Wifi Access Point Whitelist",
            ) as textDialog:
                Globals.OPEN_DIALOGS.append(textDialog)
                if textDialog.ShowModal() == wx.ID_OK:
                    Globals.OPEN_DIALOGS.remove(textDialog)
                    apList = textDialog.GetValue()
                    whitelist = []
                    parts = apList.split(",")
                    for part in parts:
                        whitelist.append(part.strip())
                    Globals.WHITELIST_AP = whitelist
                    with LargeTextEntryDialog(
                        self,
                        "Do you wish to proceed with this Wifi SSID Whitelist?",
                        "Wifi Access Point Whitelist",
                        "\n".join(Globals.WHITELIST_AP),
                        False,
                    ) as textDialog2:
                        Globals.OPEN_DIALOGS.append(textDialog2)
                        if textDialog2.ShowModal() != wx.ID_OK:
                            Globals.OPEN_DIALOGS.remove(textDialog2)
                            self.sleepInhibitor.uninhibit()
                            self.isRunning = False
                            self.setCursorDefault()
                            self.toggleEnabledState(True)
                            return
                        Globals.OPEN_DIALOGS.remove(textDialog2)
                else:
                    Globals.OPEN_DIALOGS.remove(textDialog)
                    self.sleepInhibitor.uninhibit()
                    self.isRunning = False
                    self.setCursorDefault()
                    self.toggleEnabledState(True)
                    return
        if actionClientData == GeneralActions.MOVE_GROUP.value:
            self.moveGroup()
            self.sleepInhibitor.uninhibit()
            return
        if (
            actionClientData == GeneralActions.SET_APP_STATE.value
            or actionClientData == GridActions.SET_APP_STATE.value
        ):
            self.displayAppStateChoiceDlg()
            if not self.AppState:
                self.sleepInhibitor.uninhibit()
                self.isRunning = False
                self.setCursorDefault()
                self.toggleEnabledState(True)
                return
        if actionClientData == GeneralActions.SET_DEVICE_MODE.value:
            res = None
            with wx.SingleChoiceDialog(
                self, "Select Device Mode:", "", ["Multi-App", "Kiosk"]
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
                if res == wx.ID_OK:
                    if dlg.GetStringSelection() == "Multi-App":
                        actionClientData = GeneralActions.SET_MULTI.value
                    else:
                        actionClientData = GeneralActions.SET_KIOSK.value
            if not res or res == wx.ID_CANCEL:
                self.sleepInhibitor.uninhibit()
                self.isRunning = False
                self.setCursorDefault()
                self.toggleEnabledState(True)
                return
        if (
            self.sidePanel.selectedGroupsList
            and (not self.sidePanel.selectedDevicesList or allDevicesSelected)
            and actionSelection > 0
            and actionClientData > 0
            and actionClientData < GridActions.MODIFY_ALIAS.value
        ):
            # run action on group
            if self.checkAppRequirement(actionClientData, appSelection, appLabel):
                appSelection, appLabel = self.getAppDataForRun()
            self.gridPanel.grid_1_contents = []
            self.gridPanel.grid_2_contents = []
            self.gridPanel.grid_3_contents = []
            self.gridPanel.userEdited = []
            self.gridPanel.disableGridProperties()

            groupLabel = ""
            for groupId in self.sidePanel.selectedGroupsList:
                groupLabel = list(self.sidePanel.groups.keys())[
                    list(self.sidePanel.groups.values()).index(groupId)
                ]
                self.Logging(
                    '---> Attempting to run action, "%s", on group, %s.'
                    % (actionLabel, groupLabel)
                )
            self.statusBar.gauge.Pulse()
            Globals.THREAD_POOL.enqueue(
                TakeAction, self, self.sidePanel.selectedGroupsList, actionClientData
            )
        elif (
            self.sidePanel.selectedDevicesList
            and actionSelection > 0
            and actionClientData > 0
            and actionClientData < GridActions.MODIFY_ALIAS.value
        ):
            # run action on device
            if self.checkAppRequirement(actionClientData, appSelection, appLabel):
                appSelection, appLabel = self.getAppDataForRun()
            self.gridPanel.grid_1_contents = []
            self.gridPanel.grid_2_contents = []
            self.gridPanel.grid_3_contents = []
            self.gridPanel.userEdited = []
            self.gridPanel.disableGridProperties()
            for deviceId in self.sidePanel.selectedDevicesList:
                deviceLabel = None
                try:
                    deviceLabel = list(self.sidePanel.devices.keys())[
                        list(self.sidePanel.devices.values()).index(deviceId)
                    ]
                except:
                    deviceLabel = list(self.sidePanel.devicesExtended.keys())[
                        list(self.sidePanel.devicesExtended.values()).index(deviceId)
                    ]
                self.Logging(
                    '---> Attempting to run action, "%s", on device, %s.'
                    % (actionLabel, deviceLabel)
                )
            self.statusBar.gauge.Pulse()
            Globals.THREAD_POOL.enqueue(
                TakeAction,
                self,
                self.sidePanel.selectedDevicesList,
                actionClientData,
                True,
            )
        elif actionClientData >= GridActions.MODIFY_ALIAS.value:
            # run grid action
            if self.gridPanel.grid_1.GetNumberRows() > 0:
                runAction = True
                result = None
                if Globals.SHOW_GRID_DIALOG:
                    result = CheckboxMessageBox(
                        "Confirmation",
                        "The %s will attempt to process the action on all devices in the Device Info grid.\n\nREMINDER: Only %s tags MAX may be currently applied to a device!\n\nContinue?"
                        % (Globals.TITLE, Globals.MAX_TAGS),
                    )
                    Globals.OPEN_DIALOGS.append(result)
                    if result.ShowModal() != wx.ID_OK:
                        runAction = False
                    Globals.OPEN_DIALOGS.remove(result)
                if result and result.getCheckBoxValue():
                    Globals.SHOW_GRID_DIALOG = False
                    self.preferences["gridDialog"] = Globals.SHOW_GRID_DIALOG
                if runAction:
                    if self.checkAppRequirement(
                        actionClientData, appSelection, appLabel
                    ):
                        appSelection, appLabel = self.getAppDataForRun()
                    self.Logging(
                        '---> Attempting to run grid action, "%s".' % actionLabel
                    )
                    self.gridPanel.applyTextColorToDevice(
                        None,
                        Color.black.value,
                        bgColor=Color.white.value,
                        applyAll=True,
                    )
                    self.gridPanel.disableGridProperties()
                    self.frame_toolbar.search.SetValue("")
                    self.statusBar.gauge.Pulse()
                    Globals.THREAD_POOL.enqueue(
                        iterateThroughGridRows, self, actionClientData
                    )
                else:
                    self.sleepInhibitor.uninhibit()
                    self.isRunning = False
                    self.setCursorDefault()
                    self.toggleEnabledState(True)
            else:
                displayMessageBox(
                    (
                        "Make sure the grid has data to perform the action on",
                        wx.OK | wx.ICON_ERROR,
                    )
                )
                self.sleepInhibitor.uninhibit()
                self.isRunning = False
                self.setCursorDefault()
                self.toggleEnabledState(True)
        else:
            displayMessageBox(
                (
                    "Please select an valid action to perform on the selected group(s) or device(s)!",
                    wx.OK | wx.ICON_ERROR,
                )
            )
            self.sleepInhibitor.uninhibit()
            self.isRunning = False
            self.setCursorDefault()
            self.toggleEnabledState(True)

    @api_tool_decorator()
    def checkAppRequirement(self, actionClientData, appSelection, appLabel):
        if (
            actionClientData == GeneralActions.SET_KIOSK.value
            or actionClientData == GeneralActions.CLEAR_APP_DATA.value
            or actionClientData == GeneralActions.INSTALL_APP.value
            or actionClientData == GeneralActions.UNINSTALL_APP.value
            or actionClientData == GeneralActions.SET_APP_STATE.value
            or actionClientData == GridActions.INSTALL_APP.value
            or actionClientData == GridActions.UNINSTALL_APP.value
        ) and (appSelection < 0 or appLabel == "No available app(s) on this device"):
            self.sidePanel.onAppSelection(None)
            self.sidePanel.notebook_1.SetSelection(2)
            return True
        return False

    @api_tool_decorator()
    def showConsole(self, event):
        """ Toggle Console Display """
        if not self.consoleWin:
            self.consoleWin = Console(parent=self)
            self.menubar.clearConsole.Enable(True)
            self.Bind(wx.EVT_MENU, self.onClear, self.menubar.clearConsole)
        else:
            self.consoleWin.DestroyLater()
            self.menubar.clearConsole.Enable(False)

    @api_tool_decorator()
    def onClear(self, event):
        """ Clear Console """
        if self.consoleWin:
            self.consoleWin.onClear()

    @api_tool_decorator()
    def onCommand(self, event, value="{\n\n}", level=0):
        """ When the user wants to run a command show the command dialog """
        if level < Globals.MAX_RETRY:
            self.setCursorBusy()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)

            if self.sidePanel.selectedGroupsList:
                result = None
                cmdArgs = None
                commandType = None
                schArgs = None
                schType = None
                with CommandDialog("Enter JSON Command", value=value) as cmdDialog:
                    Globals.OPEN_DIALOGS.append(cmdDialog)
                    result = cmdDialog.ShowModal()
                    Globals.OPEN_DIALOGS.remove(cmdDialog)
                    if result == wx.ID_OK:
                        try:
                            (
                                cmdArgs,
                                commandType,
                                schArgs,
                                schType,
                            ) = cmdDialog.GetValue()
                            if cmdArgs is not None:
                                createCommand(
                                    self,
                                    cmdArgs,
                                    commandType,
                                    schArgs,
                                    schType,
                                    combineRequests=True,
                                )
                        except Exception as e:
                            displayMessageBox(
                                (
                                    "An error occurred while process the inputted JSON object, please make sure it is formatted correctly",
                                    wx.OK | wx.ICON_ERROR,
                                )
                            )
                            ApiToolLog().LogError(e)
                    cmdDialog.DestroyLater()
            else:
                displayMessageBox(
                    ("Please select an group and or device", wx.OK | wx.ICON_ERROR)
                )

            self.setCursorDefault()

    @api_tool_decorator()
    def onCommandDone(self, event):
        """ Tell user to check the Esper Console for detailed results """
        cmdResult = None
        msg = ""
        if hasattr(event, "GetValue"):
            cmdResult = event.GetValue()
        else:
            cmdResult = event
        if type(cmdResult) == tuple:
            msg = cmdResult[0]
            cmdResult = cmdResult[1]
        self.menubar.enableConfigMenu()
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
        if cmdResult:
            result = ""
            for res in cmdResult:
                formattedRes = ""
                try:
                    formattedRes = json.dumps(res, indent=2).replace("\\n", "\n")
                except:
                    formattedRes = json.dumps(str(res), indent=2).replace("\\n", "\n")
                if formattedRes:
                    result += formattedRes
                    result += "\n\n"
            with ConfirmTextDialog(
                "Action has been executed.",
                "%s\n\nCheck the Esper Console for details. Last known status listed below."
                % msg
                + "\n"
                if msg
                else "",
                "Command(s) have been fired.",
                result,
                parent=self,
            ) as dialog:
                Globals.OPEN_DIALOGS.append(dialog)
                res = dialog.ShowModal()
                Globals.OPEN_DIALOGS.append(dialog)
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            (wx.CallLater, (3000, self.statusBar.setGaugeValue, 0)),
        )

    @api_tool_decorator()
    def setStatus(self, status, orgingalMsg, isError=False):
        """ Set status bar text """
        self.statusBar.sbText.SetLabel(status)
        if orgingalMsg:
            self.statusBar.sbText.SetToolTip(orgingalMsg.replace("--->", ""))
        if isError:
            self.statusBar.sbText.SetForegroundColour(Color.red.value)
        else:
            self.statusBar.sbText.SetForegroundColour(Color.black.value)

    @api_tool_decorator()
    def onFetch(self, event):
        if hasattr(self, "start_time"):
            print("Fetch Execution time: %s" % (time.time() - self.start_time))
        evtValue = event.GetValue()
        self.toggleEnabledState(False)
        if evtValue:
            action = evtValue[0]
            entId = evtValue[1]
            deviceList = evtValue[2]

            Globals.THREAD_POOL.enqueue(
                self.processFetch, action, entId, deviceList, True, len(deviceList) * 3
            )

    @api_tool_decorator()
    def processFetch(self, action, entId, deviceList, updateGauge=False, maxGauge=None):
        """ Given device data perform the specified action """
        appToUse = None
        appVersion = None
        if self.sidePanel.selectedAppEntry:
            appToUse = self.sidePanel.selectedAppEntry["pkgName"]
            appVersion = self.sidePanel.selectedAppEntry["version"]
        if (
            action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
            or action == GeneralActions.GENERATE_APP_REPORT.value
            or action == GeneralActions.GENERATE_INFO_REPORT.value
        ):
            self.gridPanel.disableGridProperties()
        num = len(deviceList)
        num = num + (num - Globals.MAX_GRID_LOAD + 1)

        isGroup = True
        if len(Globals.frame.sidePanel.selectedDevicesList) > 0:
            isGroup = False

        if action == GeneralActions.SET_KIOSK.value:
            Globals.THREAD_POOL.enqueue(
                setKiosk, self, deviceList, None, isGroup=isGroup
            )
        elif action == GeneralActions.SET_MULTI.value:
            Globals.THREAD_POOL.enqueue(
                setMulti, self, deviceList, None, isGroup=isGroup
            )
        elif action == GeneralActions.CLEAR_APP_DATA.value:
            Globals.THREAD_POOL.enqueue(clearAppData, self, deviceList, isGroup=isGroup)
        elif action == GeneralActions.SET_APP_STATE.value:
            Globals.THREAD_POOL.enqueue(
                setAppState, deviceList, appToUse, self.AppState, isGroup=isGroup
            )
        elif action == GeneralActions.REMOVE_NON_WHITELIST_AP.value:
            Globals.THREAD_POOL.enqueue(
                removeNonWhitelisted, deviceList, None, isGroup=isGroup
            )
        elif action == GeneralActions.INSTALL_APP.value:
            Globals.THREAD_POOL.enqueue(
                installAppOnDevices, appToUse, appVersion, deviceList, isGroup=isGroup
            )
        elif action == GeneralActions.UNINSTALL_APP.value:
            Globals.THREAD_POOL.enqueue(
                uninstallAppOnDevice, appToUse, deviceList, isGroup=isGroup
            )
        else:
            # populate visiable devices first
            gridDisplayData = list(deviceList.values())[: Globals.MAX_GRID_LOAD]

            # let worker threads process rest of data
            workerData = list(deviceList.values())[Globals.MAX_GRID_LOAD :]
            splitWorkerData = splitListIntoChunks(
                workerData, maxThread=((Globals.MAX_THREAD_COUNT * 2) / 3)
            )
            self.processDeviceListForGrid(
                entId, gridDisplayData, action, maxGauge, updateGauge, num
            )

            for chunk in splitWorkerData:
                Globals.THREAD_POOL.enqueue(
                    self.processDeviceListForGrid,
                    entId,
                    chunk,
                    action,
                    maxGauge,
                    updateGauge,
                    num,
                )

        Globals.THREAD_POOL.enqueue(
            self.waitForThreadsThenSetCursorDefault,
            Globals.THREAD_POOL.threads,
            3,
            action,
            tolerance=1,
        )

    def processDeviceListForGrid(
        self, entId, deviceList, action, maxGauge, updateGauge, num
    ):
        for entry in deviceList:
            if entId != Globals.enterprise_id:
                self.onClearGrids()
                break
            if checkIfCurrentThreadStopped():
                self.onClearGrids()
                break
            device = entry[0]
            deviceInfo = entry[1]

            # Populate Network sheet
            if (
                action == GeneralActions.GENERATE_INFO_REPORT.value
                or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
            ):
                if len(self.gridPanel.grid_1_contents) < Globals.MAX_GRID_LOAD + 1:
                    determineDoHereorMainThread(
                        self.gridPanel.addDeviceToNetworkGrid, device, deviceInfo
                    )
                else:
                    # construct and add info to grid contents
                    Globals.THREAD_POOL.enqueue(
                        self.gridPanel.constructNetworkGridContent, device, deviceInfo
                    )
            # Populate Device sheet
            if (
                action == GeneralActions.GENERATE_DEVICE_REPORT.value
                or action == GeneralActions.GENERATE_INFO_REPORT.value
                or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
            ):
                if len(self.gridPanel.grid_1_contents) <= Globals.MAX_GRID_LOAD + 1:
                    determineDoHereorMainThread(
                        self.gridPanel.addDeviceToDeviceGrid, deviceInfo
                    )
                else:
                    # construct and add info to grid contents
                    Globals.THREAD_POOL.enqueue(
                        self.gridPanel.constructDeviceGridContent, deviceInfo
                    )
            # Populate App sheet
            if (
                action == GeneralActions.GENERATE_APP_REPORT.value
                or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
            ):
                if len(self.gridPanel.grid_3_contents) <= Globals.MAX_GRID_LOAD + 1:
                    # To avoid adding to many app entries, add the app entry to the contents
                    # ahead of time, and then populate the grid
                    self.gridPanel.add_app_entry_to_contents(deviceInfo)
                    determineDoHereorMainThread(
                        self.gridPanel.populateAppGrid,
                        device,
                        deviceInfo,
                        deviceInfo["appObj"],
                    )
                else:
                    Globals.THREAD_POOL.enqueue(
                        self.gridPanel.constructAppGridContent,
                        device,
                        deviceInfo,
                        deviceInfo["appObj"],
                    )
            value = int(num / maxGauge * 100)
            if updateGauge and value <= 99:
                num += 1
                postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, value)

    @api_tool_decorator()
    def MacReopenApp(self, event):
        """Called when the doc icon is clicked, and ???"""
        self.onActivate(self, event, skip=False)
        if event.GetActive():
            try:
                self.GetTopWindow().Raise()
            except:
                self.tryToMakeActive()
        event.Skip()

    @api_tool_decorator()
    def MacNewFile(self):
        pass

    @api_tool_decorator()
    def MacPrintFile(self, file_path):
        pass

    @api_tool_decorator()
    def onComplete(self, event, isError=False):
        """ Things that should be done once an Action is completed """
        enable = False
        action = None
        cmdResults = None
        if event:
            eventVal = None
            if hasattr(event, "GetValue"):
                eventVal = event.GetValue()
            else:
                eventVal = event
            if type(eventVal) == tuple:
                enable = eventVal[0]
                if len(eventVal) > 1:
                    action = eventVal[1]
                if len(eventVal) > 2:
                    cmdResults = eventVal[2]
            else:
                enable = eventVal
        self.setCursorDefault()
        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
        title = "Action Completed"
        msg = ""
        if not self.IsActive() and not isError:
            if (
                self.sidePanel.actionChoice.GetClientData(
                    self.sidePanel.actionChoice.GetSelection()
                )
                == action
            ):
                actionName = self.sidePanel.actionChoice.GetValue().replace(
                    "Action -> ", ""
                )
                msg = "%s has completed." % actionName
            else:
                msg = "Action has completed."
        if self.IsFrozen():
            self.Thaw()
        self.gridPanel.button_2.Enable(self.gridArrowState["next"])
        self.gridPanel.button_1.Enable(self.gridArrowState["prev"])
        self.gridPanel.thawGridsIfFrozen()
        if self.gridPanel.disableProperties:
            self.gridPanel.enableGridProperties()
        self.gridPanel.autoSizeGridsColumns()
        if self.isRunning or enable:
            self.toggleEnabledState(True)
        self.isRunning = False
        if (
            action == GeneralActions.GENERATE_INFO_REPORT.value
            or action == GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value
        ):
            self.gridPanel.notebook_2.SetSelection(0)
        elif action == GeneralActions.GENERATE_APP_REPORT.value:
            self.gridPanel.notebook_2.SetSelection(2)
        if self.isSaving:
            self.isSaving = False
        self.sidePanel.sortAndPopulateAppChoice()
        if not self.IsIconized() and self.IsActive():
            postEventToFrame(
                eventUtil.myEVT_PROCESS_FUNCTION,
                (wx.CallLater, (3000, self.statusBar.setGaugeValue, 0)),
            )
        if cmdResults:
            postEventToFrame(
                eventUtil.myEVT_PROCESS_FUNCTION,
                (self.onCommandDone, (cmdResults,)),
            )
        self.menubar.enableConfigMenu()
        self.Logging("---> Completed Action")
        self.displayNotification(title, msg)
        self.sleepInhibitor.uninhibit()
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            self.Refresh,
        )
        if hasattr(self, "start_time"):
            print("Run Execution time: %s" % (time.time() - self.start_time))

    @api_tool_decorator()
    def onActivate(self, event, skip=True):
        if not self.isRunning and not self.isUploading and not self.isBusy:
            wx.CallLater(3000, self.statusBar.setGaugeValue, 0)
        if Globals.OPEN_DIALOGS:
            for window in Globals.OPEN_DIALOGS:
                if window and hasattr(window, "Raise") and not self.isSaving:
                    window.Raise()
                elif (
                    window and hasattr(window, "tryToMakeActive") and not self.isSaving
                ):
                    window.tryToMakeActive()
        if self.notification:
            self.notification.Close()
        self.Refresh()
        if skip:
            event.Skip()

    @api_tool_decorator()
    def onClearGrids(self, event=None):
        """ Empty Grids """
        self.gridPanel.emptyDeviceGrid()
        self.gridPanel.emptyNetworkGrid()
        self.gridPanel.emptyAppGrid()
        if event and hasattr(event, "Skip"):
            event.Skip()

    @api_tool_decorator()
    def readAuthCSV(self):
        if os.path.exists(Globals.csv_auth_path):
            if self.key and crypto().isFileEncrypt(Globals.csv_auth_path, self.key):
                crypto().decrypt(Globals.csv_auth_path, self.key, True)
            self.auth_data = read_data_from_csv_as_dict(Globals.csv_auth_path)
            if self.auth_data:
                self.auth_data = sorted(
                    self.auth_data,
                    key=lambda i: list(map(str, i["name"].lower())),
                )

    @api_tool_decorator()
    def loadPref(self):
        """ Attempt to load preferences from file system """
        if not os.path.exists(self.keyPath):
            self.key = crypto().create_key(self.keyPath)
        else:
            self.key = crypto().load_key(self.keyPath)
        if os.path.exists(self.authPath):
            Globals.csv_auth_path = self.authPath
            self.readAuthCSV()
        else:
            if self.kill:
                return
            createNewFile(self.authPath)
            write_data_to_csv(
                self.authPath, ["name", "apiHost", "enterprise", "apiKey", "apiPrefix"]
            )
            self.AddEndpoint(None)

        if self.kill:
            return

        if (
            os.path.isfile(self.prefPath)
            and os.path.exists(self.prefPath)
            and os.access(self.prefPath, os.R_OK)
        ):
            if os.path.getsize(self.prefPath) > 2:
                try:
                    self.preferences = read_json_file(self.prefPath)
                except Exception as e:
                    ApiToolLog().LogError(e)
                    self.Logging(
                        "Preference file possibly has been corrupted and is malformed!",
                        isError=True,
                    )
                self.prefDialog.SetPrefs(self.preferences)
            else:
                self.Logging("Missing or empty preference file!", isError=True)
        else:
            createNewFile(self.prefPath)
            self.savePrefs(self.prefDialog)
        self.PopulateConfig()

    @api_tool_decorator()
    def savePrefs(self, dialog):
        """ Save Preferences """
        self.preferences = dialog.GetPrefs()
        write_json_file(self.prefPath, self.preferences)
        postEventToFrame(eventUtil.myEVT_LOG, "---> Preferences' Saved")

    @api_tool_decorator()
    def onPref(self, event):
        """ Update Preferences when they are changed """
        if self.isRunning:
            return
        self.prefDialog.SetPrefs(self.preferences, onBoot=False)
        self.prefDialog.appColFilter = Globals.APP_COL_FILTER
        Globals.OPEN_DIALOGS.append(self.prefDialog)
        if self.prefDialog.ShowModal() == wx.ID_APPLY:
            self.isSavingPrefs = True
            Globals.THREAD_POOL.enqueue(self.savePrefs, self.prefDialog)
            if self.sidePanel.selectedGroupsList and self.preferences["enableDevice"]:
                self.PopulateDevices(None)
            if self.sidePanel.selectedDevicesList:
                self.sidePanel.selectedDeviceApps = []
            self.setFontSizeForLabels()
            self.handleScheduleReportPref()
        Globals.OPEN_DIALOGS.remove(self.prefDialog)
        if self.preferences and self.preferences["enableDevice"]:
            self.sidePanel.deviceChoice.Enable(True)
        else:
            self.sidePanel.deviceChoice.Enable(False)
        self.isSavingPrefs = False

    @api_tool_decorator()
    def onFail(self, event):
        """ Try to showcase rows in the grid on which an action failed on """
        failed = event.GetValue()
        if self.gridPanel.grid_1_contents and self.gridPanel.grid_2_contents:
            if type(failed) == list:
                for device in failed:
                    if "Queued" in device or "Scheduled" in device:
                        self.gridPanel.applyTextColorToDevice(
                            device[0], Color.orange.value, bgColor=Color.warnBg.value
                        )
                    else:
                        self.gridPanel.applyTextColorToDevice(
                            device, Color.red.value, bgColor=Color.errorBg.value
                        )
            elif type(failed) == tuple:
                if "Queued" in failed or "Scheduled" in failed:
                    self.gridPanel.applyTextColorToDevice(
                        failed[0], Color.orange.value, bgColor=Color.warnBg.value
                    )
                else:
                    self.gridPanel.applyTextColorToDevice(
                        failed, Color.red.value, bgColor=Color.errorBg.value
                    )
            elif failed:
                self.gridPanel.applyTextColorToDevice(
                    failed, Color.red.value, bgColor=Color.errorBg.value
                )

    @api_tool_decorator()
    def onFileDrop(self, event):
        if self.isRunning:
            return

        self.isUploading = True
        for file in event.Files:
            if file.endswith(".csv"):
                data = read_data_from_csv(file)
                if (
                    "apiHost" in data[0]
                    and "apiKey" in data[0]
                    and "apiPrefix" in data[0]
                    and "enterprise" in data[0]
                ):
                    self.PopulateConfig(auth=file)
                else:
                    if not Globals.enterprise_id:
                        displayMessageBox(
                            (
                                "Please load a configuration first!",
                                wx.OK | wx.ICON_ERROR,
                            )
                        )
                        return
                    self.toggleEnabledState(False)
                    if self.WINDOWS:
                        Globals.THREAD_POOL.enqueue(self.processDeviceCSVUpload, data)
                        Globals.THREAD_POOL.enqueue(
                            self.waitForThreadsThenSetCursorDefault,
                            Globals.THREAD_POOL.threads,
                            2,
                            tolerance=1,
                        )
                    else:
                        self.processDeviceCSVUpload(data)
                        postEventToFrame(eventUtil.myEVT_COMPLETE, True)
            elif file.endswith(".xlxs"):
                self.toggleEnabledState(False)
                if self.WINDOWS:
                    Globals.THREAD_POOL.enqueue(self.openDeviceSpreadsheet, file)
                    Globals.THREAD_POOL.enqueue(
                        self.waitForThreadsThenSetCursorDefault,
                        Globals.THREAD_POOL.threads,
                        2,
                        tolerance=1,
                    )
                else:
                    self.openDeviceSpreadsheet(file)
                    postEventToFrame(eventUtil.myEVT_COMPLETE, True)

    @api_tool_decorator()
    def onClone(self, event):
        with TemplateDialog(self.sidePanel.configChoice, parent=self) as self.tmpDialog:
            Globals.OPEN_DIALOGS.append(self.tmpDialog)
            result = self.tmpDialog.ShowModal()
            Globals.OPEN_DIALOGS.remove(self.tmpDialog)
            if result == wx.ID_OK:
                self.prepareClone(self.tmpDialog)
            self.tmpDialog.DestroyLater()

    @api_tool_decorator()
    def prepareClone(self, tmpDialog):
        self.setCursorBusy()
        self.isRunning = True
        self.statusBar.gauge.Pulse()
        util = templateUtil.EsperTemplateUtil(*tmpDialog.getInputSelections())
        Globals.THREAD_POOL.enqueue(
            util.prepareTemplate, tmpDialog.destTemplate, tmpDialog.chosenTemplate
        )

    @api_tool_decorator()
    def confirmClone(self, event):
        result = None
        res = None
        (util, toApi, toKey, toEntId, templateFound, missingApps) = event.GetValue()
        if Globals.SHOW_TEMPLATE_DIALOG:
            result = CheckboxMessageBox(
                "Confirmation",
                "The %s will attempt to clone to template.\nThe following apps are missing: %s\n\nContinue?"
                % (Globals.TITLE, missingApps if missingApps else None),
            )
            Globals.OPEN_DIALOGS.append(result)
            res = result.ShowModal()
            Globals.OPEN_DIALOGS.remove(result)
        else:
            res = wx.ID_OK
        if res == wx.ID_OK:
            Globals.THREAD_POOL.enqueue(
                self.createClone, util, templateFound, toApi, toKey, toEntId, False
            )
        else:
            self.isRunning = False
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
            self.setCursorDefault()
        if result and result.getCheckBoxValue():
            Globals.SHOW_TEMPLATE_DIALOG = False
            self.preferences["templateDialog"] = Globals.SHOW_TEMPLATE_DIALOG

    @api_tool_decorator()
    def confirmCloneUpdate(self, event):
        result = None
        res = None
        (util, toApi, toKey, toEntId, templateFound, missingApps) = event.GetValue()
        if Globals.SHOW_TEMPLATE_UPDATE:
            result = CheckboxMessageBox(
                "Confirmation",
                "The Template already exists on the destination tenant.\nThe following apps are missing: %s\n\nWould you like to update the template?"
                % (missingApps if missingApps else None),
            )
            Globals.OPEN_DIALOGS.append(result)
            res = result.ShowModal()
            Globals.OPEN_DIALOGS.remove(result)
        else:
            res = wx.ID_OK
        if res == wx.ID_OK:
            Globals.THREAD_POOL.enqueue(
                self.createClone, util, templateFound, toApi, toKey, toEntId, True
            )
        else:
            self.isRunning = False
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
            self.setCursorDefault()
        if result and result.getCheckBoxValue():
            Globals.SHOW_TEMPLATE_UPDATE = False
            self.preferences["templateUpdate"] = Globals.SHOW_TEMPLATE_UPDATE

    @api_tool_decorator()
    def createClone(
        self, util, templateFound, toApi, toKey, toEntId, update=False, level=0
    ):
        if level == 0:
            templateFound = util.processDeviceGroup(templateFound)
            templateFound = util.processWallpapers(templateFound)
        self.Logging("Attempting to copy template...")
        res = None
        try:
            if update:
                res = util.updateTemplate(toApi, toKey, toEntId, templateFound)
            else:
                res = util.createTemplate(
                    toApi, toKey, toEntId, templateFound, level + 1
                )
        except Exception as e:
            res = e
        if type(res) == dict and "errors" not in res:
            action = "created" if not update else "updated"
            self.Logging("Template sucessfully %s." % action)
            displayMessageBox(
                ("Template sucessfully %s." % action, wx.OK | wx.ICON_INFORMATION)
            )
        elif (
            type(res) == dict
            and "errors" in res
            and res["errors"]
            and "EMM" in res["errors"][0]
            and level < 2
        ) or (isinstance(res, Exception) and "EMM" in str(res)):
            del templateFound["template"]["application"]["managed_google_play_disabled"]
            self.createClone(
                util, templateFound, toApi, toKey, toEntId, update, level + 1
            )
        else:
            action = "recreate" if not update else "update"
            self.Logging("ERROR: Failed to %s Template.%s" % (action, res))
            displayMessageBox(
                (
                    "ERROR: Failed to %s Template. Please try again." % action,
                    wx.OK | wx.ICON_ERROR,
                )
            )
        self.isRunning = False
        postEventToFrame(eventUtil.myEVT_COMPLETE, True)

    @api_tool_decorator()
    def onSearch(self, event=None):
        queryString = ""
        if hasattr(event, "GetString"):
            queryString = event.GetString()
        elif isinstance(event, str):
            queryString = event
        else:
            queryString = self.frame_toolbar.search.GetValue()

        if self.searchThreads:
            for thread in self.searchThreads:
                if thread.is_alive():
                    if hasattr(thread, "stop"):
                        thread.stop()
            self.searchThreads = []

        t = wxThread.GUIThread(
            self,
            self.processSearch,
            args=(event, queryString),
            name="processSearch",
        )
        t.startWithRetry()
        self.searchThreads.append(t)

    def processSearch(self, event, queryString):
        if self.searchThreads:
            for thread in self.searchThreads:
                if thread.is_alive() and threading.current_thread() != thread:
                    thread.join()

        self.setCursorBusy()
        self.gridPanel.setGridsCursor(wx.Cursor(wx.CURSOR_WAIT))
        self.gridPanel.disableGridProperties()
        if (
            hasattr(event, "EventType")
            and (
                (wx.EVT_TEXT.typeId == event.EventType and not queryString)
                or (wx.EVT_SEARCH.typeId == event.EventType and not queryString)
                or wx.EVT_SEARCH_CANCEL.typeId == event.EventType
            )
            or wx.EVT_CHAR.typeId == event
        ):
            postEventToFrame(
                eventUtil.myEVT_PROCESS_FUNCTION,
                (self.applySearchColor, (queryString, Color.white.value, True)),
            )
        if queryString:
            postEventToFrame(
                eventUtil.myEVT_PROCESS_FUNCTION,
                (self.applySearchColor, (queryString, Color.lightYellow.value)),
            )
            self.Logging("--> Search for %s completed" % queryString)
        else:
            self.frame_toolbar.search.SetValue("")
            postEventToFrame(
                eventUtil.myEVT_PROCESS_FUNCTION,
                (self.applySearchColor, (queryString, Color.white.value, True)),
            )
        self.setCursorDefault()
        self.gridPanel.setGridsCursor(wx.Cursor(wx.CURSOR_DEFAULT))
        self.gridPanel.enableGridProperties()

    def applySearchColor(self, queryString, color, applyAll=False):
        self.gridPanel.applyTextColorMatchingGridRow(
            self.gridPanel.grid_1, queryString, color, applyAll
        )
        self.gridPanel.applyTextColorMatchingGridRow(
            self.gridPanel.grid_2, queryString, color, applyAll
        )
        self.gridPanel.applyTextColorMatchingGridRow(
            self.gridPanel.grid_3, queryString, color, applyAll
        )

    @api_tool_decorator()
    def toggleEnabledState(self, state):
        determineDoHereorMainThread(self.sidePanel.runBtn.Enable, state)
        determineDoHereorMainThread(self.sidePanel.actionChoice.Enable, state)
        determineDoHereorMainThread(self.sidePanel.removeEndpointBtn.Enable, state)

        if not self.sidePanel.appChoice.IsEnabled() and state:
            action = self.sidePanel.actionChoice.GetValue()
            clientData = None
            if action in Globals.GENERAL_ACTIONS:
                clientData = Globals.GENERAL_ACTIONS[action]
            elif action in Globals.GRID_ACTIONS:
                clientData = Globals.GRID_ACTIONS[action]
            determineDoHereorMainThread(self.sidePanel.setAppChoiceState, clientData)
        else:
            determineDoHereorMainThread(self.sidePanel.appChoice.Enable, state)

        determineDoHereorMainThread(
            self.frame_toolbar.EnableTool, self.frame_toolbar.otool.Id, state
        )
        determineDoHereorMainThread(
            self.frame_toolbar.EnableTool, self.frame_toolbar.rtool.Id, state
        )
        determineDoHereorMainThread(
            self.frame_toolbar.EnableTool, self.frame_toolbar.cmdtool.Id, state
        )
        determineDoHereorMainThread(
            self.frame_toolbar.EnableTool, self.frame_toolbar.atool.Id, state
        )

        # Toggle Menu Bar Items
        determineDoHereorMainThread(self.menubar.fileOpenAuth.Enable, state)
        for option in self.menubar.sensitiveMenuOptions:
            determineDoHereorMainThread(self.menubar.EnableTop, option, state)
        determineDoHereorMainThread(self.menubar.fileOpenConfig.Enable, state)
        determineDoHereorMainThread(self.menubar.pref.Enable, state)
        determineDoHereorMainThread(self.menubar.collection.Enable, state)
        determineDoHereorMainThread(self.menubar.eqlQuery.Enable, state)
        determineDoHereorMainThread(self.menubar.run.Enable, state)
        determineDoHereorMainThread(self.menubar.installedDevices.Enable, state)
        determineDoHereorMainThread(self.menubar.command.Enable, state)
        determineDoHereorMainThread(self.menubar.collectionSubMenu.Enable, state)
        determineDoHereorMainThread(self.menubar.groupSubMenu.Enable, state)
        determineDoHereorMainThread(self.menubar.setSaveMenuOptionsEnableState, state)

        if not self.blueprintsEnabled:
            determineDoHereorMainThread(self.menubar.clone.Enable, state)
        else:
            determineDoHereorMainThread(self.menubar.cloneBP.Enable, state)

    @api_tool_decorator()
    def onInstalledDevices(self, event):
        reset = True
        if self.sidePanel.apps:
            self.setCursorBusy()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
            self.toggleEnabledState(False)
            with InstalledDevicesDlg(self.sidePanel.apps) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
                if res == wx.ID_OK:
                    app, version = dlg.getAppValues()
                    if app and version:
                        defaultFileName = "%s_%s_installed_devices.xlsx" % (
                            dlg.selectedAppName.strip().replace(" ", "-").lower(),
                            str(dlg.selectedVersion) if not "All" in dlg.selectedVersion else "all-versions",
                        )
                        inFile = displaySaveDialog(
                            "Save Installed Devices to CSV",
                            "Microsoft Excel Open XML Spreadsheet (*.xlsx)|*.xlsx|CSV files (*.csv)|*.csv",
                            defaultFile=defaultFileName,
                        )
                        if inFile:
                            self.onClearGrids()
                            self.statusBar.gauge.Pulse()
                            self.setCursorBusy()
                            self.isRunning = True
                            self.toggleEnabledState(False)
                            self.sleepInhibitor.inhibit()
                            reset = False
                            Globals.THREAD_POOL.enqueue(
                                fetchInstalledDevices, app, version, inFile
                            )
                        else:
                            self.setCursorDefault()
                            self.toggleEnabledState(True)
                    else:
                        displayMessageBox(
                            (
                                "Failed to get app id or version",
                                wx.ICON_INFORMATION,
                            )
                        )
                        postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE_LATER, (3000, 0))
                        self.setCursorDefault()
                        self.toggleEnabledState(True)
                dlg.DestroyLater()
        else:
            displayMessageBox(
                (
                    "No apps found",
                    wx.ICON_INFORMATION,
                )
            )
        if reset:
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE_LATER, (3000, 0))
            self.setCursorDefault()
            self.toggleEnabledState(True)

    @api_tool_decorator()
    def moveGroup(self, event=None):
        if self.sidePanel.selectedDevicesList:
            choices = list(self.sidePanel.groups.keys())
            groupMultiDialog = MultiSelectSearchDlg(
                self,
                choices,
                label="Select Group to move to",
                title="Select Group to move to",
                single=True,
                resp=self.sidePanel.groupsResp,
            )
            Globals.OPEN_DIALOGS.append(groupMultiDialog)
            if groupMultiDialog.ShowModal() == wx.ID_OK:
                Globals.OPEN_DIALOGS.remove(groupMultiDialog)
                selections = groupMultiDialog.GetSelections()
                if selections:
                    selction = selections[0]
                    groups = getAllGroups(name=selction)
                if groups.results:
                    resp = moveGroup(
                        groups.results[0].id, self.sidePanel.selectedDevicesList
                    )
                    if resp and resp.status_code == 200:
                        displayMessageBox(
                            "Selected device(s) have been moved to the %s Group."
                            % groups.results[0].name
                        )
                        self.sidePanel.clearSelections()
                    elif resp:
                        displayMessageBox(str(resp))
                else:
                    displayMessageBox("No Group found with the name: %s" % selction)
                postEventToFrame(eventUtil.myEVT_COMPLETE, True)
                return
            else:
                Globals.OPEN_DIALOGS.remove(groupMultiDialog)
                self.isRunning = False
                self.setCursorDefault()
                self.toggleEnabledState(True)
                return
        else:
            displayMessageBox(
                (
                    "Please select a Group and then the device(s) you wish to move from the selectors!",
                    wx.OK | wx.ICON_ERROR,
                )
            )
            self.isRunning = False
            self.setCursorDefault()
            self.toggleEnabledState(True)

    @api_tool_decorator()
    def createGroup(self, event):
        if not self.groupManage:
            self.groupManage = GroupManagement(self.groups)
        with self.groupManage as manage:
            Globals.OPEN_DIALOGS.append(manage)
            manage.ShowModal()
            Globals.OPEN_DIALOGS.remove(manage)

    @api_tool_decorator()
    def installApp(self, event):
        if self.sidePanel.selectedGroupsList or self.sidePanel.selectedDevicesList:
            res = version = pkg = None
            with InstalledDevicesDlg(
                self.sidePanel.enterpriseApps,
                title="Install Application",
                showAllVersionsOption=False,
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
                if res == wx.ID_OK:
                    _, version, pkg = dlg.getAppValues(returnPkgName=True)
            if pkg:
                if self.sidePanel.selectedDevicesList:
                    Globals.THREAD_POOL.enqueue(
                        installAppOnDevices, pkg, version, postStatus=True
                    )
                elif self.sidePanel.selectedGroupsList:
                    Globals.THREAD_POOL.enqueue(
                        installAppOnGroups, pkg, version, postStatus=True
                    )
        else:
            displayMessageBox(
                (
                    "Please select the group(s) and or device(s) you wish to install an app to!",
                    wx.OK | wx.ICON_ERROR,
                )
            )

    @api_tool_decorator()
    def uninstallApp(self, event):
        if self.sidePanel.selectedGroupsList or self.sidePanel.selectedDevicesList:
            res = pkg = None
            with InstalledDevicesDlg(
                self.sidePanel.apps,
                hide_version=True,
                title="Uninstall Application",
                showAllVersionsOption=False,
                showPkgTextInput=True,
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
            if res == wx.ID_OK:
                _, _, pkg = dlg.getAppValues(returnPkgName=True)
            if pkg:
                if self.sidePanel.selectedDevicesList:
                    Globals.THREAD_POOL.enqueue(
                        uninstallAppOnDevice, pkg, postStatus=True
                    )
                elif self.sidePanel.selectedGroupsList:
                    Globals.THREAD_POOL.enqueue(
                        uninstallAppOnGroup, pkg, postStatus=True
                    )
        else:
            displayMessageBox(
                (
                    "Please select the group(s) and or device(s) you wish to uninstall an app from!",
                    wx.OK | wx.ICON_ERROR,
                )
            )

    @api_tool_decorator()
    def callSetGaugeLater(self, event):
        delayMs = 3000
        value = 0
        if event and hasattr(event, "GetValue"):
            val = event.GetValue()
            if type(val) == tuple:
                delayMs = val[0]
                value = val[1]
        wx.CallLater(delayMs, self.statusBar.setGaugeValue, value)

    @api_tool_decorator()
    def displayNotificationEvent(self, event):
        title = msg = ""
        if event and hasattr(event, "GetValue"):
            val = event.GetValue()
            if type(val) == tuple:
                title = val[0]
                msg = val[1]
        self.displayNotification(title, msg)

    @api_tool_decorator()
    def displayNotification(self, title, msg, displayActive=False):
        if not self.IsActive() or displayActive:
            self.notification = wxadv.NotificationMessage(title, msg, self)
            if self.notification:
                if hasattr(self.notification, "MSWUseToasts"):
                    try:
                        self.notification.MSWUseToasts()
                    except:
                        pass
                self.notification.Show()

    @api_tool_decorator()
    def onSuspend(self, event):
        if (
            self.isRunning
            or self.isRunningUpdate
            or self.isSavingPrefs
            or self.isUploading
            or self.isBusy
            and hasattr(event, "Veto")
        ):
            event.Veto()

    @api_tool_decorator()
    def displayAppStateChoiceDlg(self):
        res = None
        with wx.SingleChoiceDialog(
            self, "Select App State:", "", ["DISABLE", "HIDE", "SHOW"]
        ) as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            res = dlg.ShowModal()
            Globals.OPEN_DIALOGS.remove(dlg)
            if res == wx.ID_OK:
                self.AppState = dlg.GetStringSelection()
            else:
                self.AppState = None

    @api_tool_decorator()
    def uploadApplication(self, event=None, title="", joinThread=False):
        with wx.FileDialog(
            self,
            "Upload APK" if not title else title,
            wildcard="APK files (*.apk)|*.apk",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
            defaultDir=str(self.defaultDir),
        ) as fileDialog:
            Globals.OPEN_DIALOGS.append(fileDialog)
            result = fileDialog.ShowModal()
            Globals.OPEN_DIALOGS.remove(fileDialog)
            if result == wx.ID_OK:
                apk_path = fileDialog.GetPath()
                Globals.THREAD_POOL.enqueue(uploadAppToEndpoint, apk_path)
                if joinThread:
                    Globals.THREAD_POOL.join()
        if event:
            event.Skip()

    @api_tool_decorator()
    def onBulkFactoryReset(self, event):
        with BulkFactoryReset() as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            res = dlg.ShowModal()
            Globals.OPEN_DIALOGS.remove(dlg)

            if res == wx.ID_OK:
                self.statusBar.gauge.Pulse()
                ids = dlg.getIdentifiers()
                bulkFactoryReset(ids)

    @api_tool_decorator()
    def onGeofence(self, event):
        with GeofenceDialog() as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            dlg.ShowModal()
            Globals.OPEN_DIALOGS.remove(dlg)

    @api_tool_decorator()
    def onCloneBP(self, event):
        with BlueprintsDialog(self.sidePanel.configChoice, parent=self) as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            result = dlg.ShowModal()
            Globals.OPEN_DIALOGS.remove(dlg)
            if result == wx.ID_OK:
                prepareBlueprintClone(
                    dlg.getBlueprint(),
                    dlg.toConfig,
                    dlg.fromConfig,
                    dlg.getDestinationGroup(),
                )
                pass
            dlg.DestroyLater()

    @api_tool_decorator()
    def loadConfigCheckBlueprint(self, config):
        Globals.token_lock.acquire()
        Globals.token_lock.release()
        if "isBlueprintsEnabled" in config:
            self.blueprintsEnabled = config["isBlueprintsEnabled"]
        else:
            checkBlueprintEnabled(config)
            self.blueprintsEnabled = config["isBlueprintsEnabled"]
        if self.blueprintsEnabled:
            self.menubar.toggleCloneMenuOptions(True)
        else:
            self.menubar.toggleCloneMenuOptions(False)

    @api_tool_decorator()
    def onUserReport(self, event):
        defaultFileName = "%s_User-Report" % datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )
        self.isSaving = True
        inFile = displaySaveDialog(
            "Save User Report as...",
            "CSV files (*.csv)|*.csv",
            defaultFile=defaultFileName,
        )

        if inFile:  # Save button was pressed
            self.statusBar.gauge.Pulse()
            self.setCursorBusy()
            self.toggleEnabledState(False)
            self.gridPanel.disableGridProperties()
            users = getAllUsers()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 50)
            data = [
                [
                    "User Id",
                    "Username",
                    "Email",
                    "First Name",
                    "Last Name",
                    "Full Name",
                    "Is Active",
                    "Role",
                    "Groups",
                    "Created On",
                    "Updated On",
                    "Last Login",
                ]
            ]
            num = 1
            for user in users["results"]:
                entry = []
                entry.append(user["id"])
                entry.append(user["username"])
                entry.append(user["email"])
                entry.append(user["first_name"])
                entry.append(user["last_name"])
                entry.append(user["full_name"])
                entry.append(user["is_active"])
                entry.append(user["profile"]["role"])
                entry.append(user["profile"]["groups"])
                entry.append(user["profile"]["created_on"])
                entry.append(user["profile"]["updated_on"])
                entry.append(user["last_login"])
                data.append(entry)
                postEventToFrame(
                    eventUtil.myEVT_UPDATE_GAUGE, int(num / len(users["results"]) * 90)
                )
                num += 1
            createNewFile(inFile)

            write_data_to_csv(inFile, data)

            self.Logging("---> User Report saved to file: " + inFile)
            self.setCursorDefault()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
            self.toggleEnabledState(True)
            self.gridPanel.enableGridProperties()

            res = displayMessageBox(
                (
                    "User Report Saved\n\n File saved at: %s\n\nWould you like to navigate to the file?"
                    % inFile,
                    wx.YES_NO | wx.ICON_INFORMATION,
                )
            )
            self.isSaving = False
            if res == wx.YES:
                parentDirectory = Path(inFile).parent.absolute()
                openWebLinkInBrowser(parentDirectory, isfile=True)

    @api_tool_decorator()
    def onPendingUserReport(self, event):
        defaultFileName = "%s_Pending-User-Report" % datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )
        self.isSaving = True
        inFile = displaySaveDialog(
            "Save Pending User Report as...",
            "CSV files (*.csv)|*.csv",
            defaultFile=defaultFileName,
        )

        if inFile:  # Save button was pressed
            self.statusBar.gauge.Pulse()
            self.setCursorBusy()
            self.toggleEnabledState(False)
            self.gridPanel.disableGridProperties()
            users = getAllPendingUsers()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 50)
            data = [
                [
                    "User Id",
                    "Email",
                    "Is Active",
                    "Role",
                    "Groups",
                    "Created On",
                    "Updated On",
                ]
            ]
            num = 1
            for user in users["userinvites"]:
                entry = []
                entry.append(user["id"])
                entry.append(user["email"])
                entry.append(user["meta"]["is_active"])
                entry.append(user["meta"]["profile"]["role"])
                entry.append(user["meta"]["profile"]["groups"])
                entry.append(user["created_at"])
                entry.append(user["updated_at"])
                data.append(entry)
                postEventToFrame(
                    eventUtil.myEVT_UPDATE_GAUGE,
                    int(num / len(users["userinvites"]) * 90),
                )
                num += 1
            createNewFile(inFile)

            write_data_to_csv(inFile, data)

            self.Logging("---> Pending User Report saved to file: " + inFile)
            self.setCursorDefault()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
            self.toggleEnabledState(True)
            self.gridPanel.enableGridProperties()

            res = displayMessageBox(
                (
                    "Pending User Report Saved\n\n File saved at: %s\n\nWould you like to navigate to the file?"
                    % inFile,
                    wx.YES_NO | wx.ICON_INFORMATION,
                )
            )
            self.isSaving = False
            if res == wx.YES:
                parentDirectory = Path(inFile).parent.absolute()
                openWebLinkInBrowser(parentDirectory, isfile=True)

    @api_tool_decorator()
    def onConvertTemplate(self, event):
        with BlueprintsConvertDialog(self.sidePanel.configChoice, parent=self) as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            result = dlg.ShowModal()
            Globals.OPEN_DIALOGS.remove(dlg)
            if result == wx.ID_OK:
                prepareBlueprintConversion(
                    dlg.getTemplate(),
                    dlg.toConfig,
                    dlg.fromConfig,
                    dlg.getDestinationGroup(),
                )
                pass
            dlg.DestroyLater()

    @api_tool_decorator()
    def downloadGroups(self, event):
        defaultFileName = "%s_Group-Report" % datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )
        self.isSaving = True
        inFile = displaySaveDialog(
            "Save Group Report as...",
            "CSV files (*.csv)|*.csv",
            defaultFile=defaultFileName,
        )

        if inFile:  # Save button was pressed
            self.statusBar.gauge.Pulse()
            self.setCursorBusy()
            self.toggleEnabledState(False)
            self.gridPanel.disableGridProperties()

            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 50)
            data = [
                [
                    "Group Id",
                    "Group Name",
                    "Group Path",
                    "Description",
                    "Device Count",
                    "Count of Child Groups",
                    "Parent Group URL",
                    "Thumbnail",
                    "Blueprint",
                    "Created On",
                    "Updated On",
                ]
            ]
            num = 1
            for group in self.sidePanel.groupsResp["results"]:
                entry = []
                entry.append(group["id"])
                entry.append(group["name"])
                entry.append(group["path"])
                entry.append(group["description"])
                entry.append(group["device_count"])
                entry.append(group["children_count"])
                entry.append(group["parent"] if "parent" in group else "")
                entry.append(group["thumbnail"])
                entry.append(group["blueprint"])
                entry.append(group["created_on"])
                entry.append(group["updated_on"])
                data.append(entry)
                postEventToFrame(
                    eventUtil.myEVT_UPDATE_GAUGE,
                    int(num / len(self.sidePanel.groupsResp["results"]) * 90),
                )
                num += 1
            createNewFile(inFile)

            write_data_to_csv(inFile, data)

            self.Logging("---> User Report saved to file: " + inFile)
            self.setCursorDefault()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 100)
            self.toggleEnabledState(True)
            self.gridPanel.enableGridProperties()

            res = displayMessageBox(
                (
                    "Group Report Saved\n\n File saved at: %s\n\nWould you like to navigate to the file?"
                    % inFile,
                    wx.YES_NO | wx.ICON_INFORMATION,
                )
            )
            self.isSaving = False
            if res == wx.YES:
                parentDirectory = Path(inFile).parent.absolute()
                openWebLinkInBrowser(parentDirectory, isfile=True)

    def onNewBlueprintApp(self, event):
        reset = True
        if self.sidePanel.apps:
            self.setCursorBusy()
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE, 0)
            self.toggleEnabledState(False)
            with InstalledDevicesDlg(
                self.sidePanel.apps,
                title="Select New Blueprint App",
                showAllVersionsOption=False,
                showBlueprintInput=True,
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
                if res == wx.ID_OK:
                    selection, apps = dlg.getBlueprintInputs()
                    self.Logging("Fetching List of Blueprints...")
                    blueprints = getAllBlueprints()
                    self.Logging("Processing List of Blueprints...")
                    Globals.THREAD_POOL.enqueue(
                        modifyAppsInBlueprints,
                        blueprints,
                        apps,
                        self.changedBlueprints,
                        addToAppListIfNotPresent=selection,
                    )
        else:
            displayMessageBox(
                (
                    "No apps found",
                    wx.ICON_INFORMATION,
                )
            )
        if reset:
            postEventToFrame(eventUtil.myEVT_UPDATE_GAUGE_LATER, (3000, 0))
            self.setCursorDefault()
            self.toggleEnabledState(True)

    def displayBlueprintActionDlg(self, success, total):
        res = None
        opt = None
        options = ["Apply Immediatly", "Schedule", "Do Nothing"]
        if success > 0:
            with wx.SingleChoiceDialog(
                self,
                "Successfully changed %s of %s Blueprints.\nWhat action would you like to apply to the changed Blueprints?"
                % (success, total),
                "",
                options,
            ) as dlg:
                Globals.OPEN_DIALOGS.append(dlg)
                res = dlg.ShowModal()
                Globals.OPEN_DIALOGS.remove(dlg)
                if res == wx.ID_OK:
                    opt = dlg.GetStringSelection()
                else:
                    return None

            statusList = []
            if not opt or opt == options[2]:
                return None
            elif opt == options[0]:
                # Push immediately
                num = 1
                for bp in self.changedBlueprints:
                    updateResp, _ = pushBlueprintUpdate(bp["id"], bp["group"])
                    statusList.append(
                        {
                            "Blueprint Id": bp["id"],
                            "Blueprint Name": bp["name"],
                            "Group Id": bp["group"],
                            "Group Path": Globals.knownGroups[bp["group"]]["path"]
                            if bp["group"] in Globals.knownGroups
                            else "Unknown",
                            "Response": updateResp.text,
                        }
                    )
                    postEventToFrame(
                        eventUtil.myEVT_UPDATE_GAUGE,
                        int(num / len(self.changedBlueprints) * 100),
                    )
                    num += 1
            elif opt == options[1]:
                # prompt for schedule
                schedule = None
                scheduleType = "WINDOW"
                with ScheduleCmdDialog() as dlg:
                    Globals.OPEN_DIALOGS.append(dlg)
                    res = dlg.ShowModal()
                    Globals.OPEN_DIALOGS.remove(dlg)
                    if res == wx.ID_OK:
                        if dlg.isRecurring():
                            scheduleType = "RECURRING"
                        schedule = dlg.getScheduleDict()
                    else:
                        return None

                num = 1
                for bp in self.changedBlueprints:
                    updateResp, _ = pushBlueprintUpdate(
                        bp["id"],
                        bp["group"],
                        schedule=schedule,
                        schedule_type=scheduleType,
                    )
                    statusList.append(
                        {
                            "Blueprint Id": bp["id"],
                            "Blueprint Name": bp["name"],
                            "Group Id": bp["group"],
                            "Group Path": Globals.knownGroups[bp["group"]]["path"]
                            if bp["group"] in Globals.knownGroups
                            else "Unknown",
                            "Response": updateResp.text,
                        }
                    )
                    postEventToFrame(
                        eventUtil.myEVT_UPDATE_GAUGE,
                        int(num / len(self.changedBlueprints) * 100),
                    )
                    num += 1

            postEventToFrame(eventUtil.myEVT_COMMAND, statusList)
        else:
            displayMessageBox(
                (
                    "Successfully changed %s of %s Blueprints." & (success, total),
                    wx.ICON_INFORMATION,
                )
            )

    def handleScheduleReportPref(self):
        if Globals.SCHEDULE_ENABLED:
            if self.scheduleReport:
                # Stop exisiting report, just in case timing differs
                self.scheduleReport.stop()
                self.stopOtherScheduledCalls()
            # Start scheduled report
            self.scheduleReport = wxThread.GUIThread(
                self, self.beginScheduledReport, None, name="ScheduledReportThread"
            )
            self.scheduleReport.startWithRetry()
        elif self.scheduleReport:
            # Stop report thread as it should be disabled
            self.scheduleReport.stop()
            self.stopOtherScheduledCalls()

    def beginScheduledReport(self):
        if not Globals.SCHEDULE_ENABLED or checkIfCurrentThreadStopped():
            return

        dirPath = Globals.SCHEDULE_LOCATION
        if not os.path.exists(dirPath):
            os.mkdir(dirPath)
        fileName = "%s_EAST-Report.%s" % (
            datetime.now().strftime("%Y-%m-%d_%H-%M-%S"),
            Globals.SCHEDULE_SAVE,
        )
        filePath = os.path.join(dirPath, fileName)

        if self.delayScheduleReport:
            time.sleep(60 * 5)
        self.delayScheduleReport = False

        self.waitUntilNotBusy()

        if not Globals.SCHEDULE_ENABLED or checkIfCurrentThreadStopped():
            return

        self.Logging("Performing scheduled report")
        correctSaveFileName(filePath)
        self.setCursorBusy()
        self.toggleEnabledState(False)
        self.gridPanel.disableGridProperties()
        self.Logging("Attempting to save scheduled report at %s" % filePath)
        self.statusBar.gauge.Pulse()

        reportAction = GeneralActions.GENERATE_INFO_REPORT.value
        if Globals.SCHEDULE_TYPE == "Device":
            reportAction = GeneralActions.GENERATE_DEVICE_REPORT.value
        elif Globals.SCHEDULE_TYPE == "App":
            reportAction = GeneralActions.GENERATE_APP_REPORT.value
        elif Globals.SCHEDULE_TYPE == "All":
            reportAction = GeneralActions.SHOW_ALL_AND_GENERATE_REPORT.value

        self.scheduleReportRunning = True
        Globals.THREAD_POOL.enqueue(
            self.saveAllFile,
            filePath,
            action=reportAction,
            showDlg=False,
            allDevices=True,
            tolarance=1,
        )
        Globals.THREAD_POOL.join()

        # Schedule next occurrance of report
        self.processScheduleCallLater()
        self.scheduleReportRunning = False

    def processScheduleCallLater(self):
        self.stopOtherScheduledCalls()
        postEventToFrame(
            eventUtil.myEVT_PROCESS_FUNCTION,
            (self.startScheduleReportCall),
        )

    def startScheduleReportCall(self):
        self.scheduleCallLater.append(
            wx.CallLater(
                Globals.SCHEDULE_INTERVAL * 3600000, self.handleScheduleReportPref
            )
        )

    def stopOtherScheduledCalls(self):
        for entry in self.scheduleCallLater:
            if hasattr(entry, "Stop"):
                entry.Stop()
                wx.CallAfter(entry.Notify)
        self.scheduleCallLater = []

    def waitUntilNotBusy(self, amountSleep=60):
        # Pause until a minute after current task is complete
        while (
            self.isRunning
            or self.scheduleReportRunning
            or self.isRunningUpdate
            or self.isSavingPrefs
            or self.isUploading
            or self.isBusy
            or self.isSaving
            or Globals.OPEN_DIALOGS
        ):
            time.sleep(amountSleep)

    @api_tool_decorator()
    def onConfigureWidgets(self, event):
        res = None
        enable = className = deviceList = None
        with WidgetPicker() as dlg:
            Globals.OPEN_DIALOGS.append(dlg)
            res = dlg.ShowModal()
            enable, className, commandTarget, deviceList = dlg.getInputs()
            Globals.OPEN_DIALOGS.remove(dlg)
        if res == wx.ID_APPLY:
            self.statusBar.gauge.Pulse()
            if commandTarget == 0:
                self.Logging("Creating Widget Command for selected devices")
                Globals.THREAD_POOL.enqueue(
                    setWidget, enable, widgetName=className, devices=deviceList
                )
            else:
                self.Logging("Creating Widget Command for selected groups")
                Globals.THREAD_POOL.enqueue(
                    setWidget, enable, widgetName=className, groups=deviceList
                )
